"""
Dynamic PostgreSQL Explorer & Report Generator
================================================
A production-ready Streamlit application that lets a user pick any table
in a PostgreSQL database, filter it by a reference column (date/timestamp
or categorical), pick one or more output columns, and view/download the
resulting report with auto-generated KPI cards for numeric columns.

Run with:
    streamlit run app.py
Configuration is read from environment variables (see .env.example / README).
"""

from __future__ import annotations

import io
import math
import os
import re
from datetime import date
from typing import Any

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, inspect, text
from sqlalchemy import types as sa_types
from sqlalchemy.engine import Engine
from sqlalchemy.exc import SQLAlchemyError

# --------------------------------------------------------------------------- #
# Page configuration
# --------------------------------------------------------------------------- #
st.set_page_config(
    page_title="DB Explorer & Report Generator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# --------------------------------------------------------------------------- #
# Configuration / Environment
# --------------------------------------------------------------------------- #
# Priority: st.secrets (if running on Streamlit Cloud / has a secrets.toml)
# falls back to plain environment variables. This keeps local dev (.env via
# os.environ, e.g. loaded through `python-dotenv` or exported in shell) and
# cloud deployment (st.secrets) both working without code changes.


def _get_config_value(key: str, default: str | None = None) -> str | None:
    """Fetch a config value from st.secrets first, then environment vars."""
    try:
        if key in st.secrets:
            return str(st.secrets[key])
    except Exception:
        # st.secrets raises if no secrets.toml exists at all -- that's fine,
        # we just fall through to environment variables.
        pass
    return os.environ.get(key, default)


DB_HOST = _get_config_value("DB_HOST", "localhost")
DB_PORT = _get_config_value("DB_PORT", "5432")
DB_NAME = _get_config_value("DB_NAME", "postgres")
DB_USER = _get_config_value("DB_USER", "postgres")
DB_PASSWORD = _get_config_value("DB_PASSWORD", "")
DB_SSLMODE = _get_config_value("DB_SSLMODE", "prefer")  # e.g. "require" for cloud DBs

# Allow a full DATABASE_URL to override the individual pieces above.
DATABASE_URL = _get_config_value("DATABASE_URL")


# --------------------------------------------------------------------------- #
# Database Engine (cached across reruns / sessions)
# --------------------------------------------------------------------------- #
@st.cache_resource(show_spinner="Connecting to database...")
def get_engine() -> Engine:
    """
    Build and cache a SQLAlchemy engine for the app's lifetime.

    Using @st.cache_resource ensures a single connection pool is reused
    across reruns instead of opening a new connection on every interaction.
    """
    if DATABASE_URL:
        url = DATABASE_URL
    else:
        url = (
            f"postgresql+psycopg2://{DB_USER}:{DB_PASSWORD}"
            f"@{DB_HOST}:{DB_PORT}/{DB_NAME}?sslmode={DB_SSLMODE}"
        )

    engine = create_engine(
        url,
        pool_pre_ping=True,   # detect dead connections and recycle them
        pool_size=5,
        max_overflow=5,
        pool_recycle=1800,    # recycle connections every 30 minutes
    )
    # Fail fast if credentials/host are wrong, so we can show a clean error
    # at startup rather than on the first query the user runs.
    with engine.connect() as conn:
        conn.execute(text("SELECT 1"))
    return engine


def get_engine_safe() -> Engine | None:
    """Wrapper that turns connection failures into a clean st.error message."""
    try:
        return get_engine()
    except SQLAlchemyError as exc:
        st.error(
            "❌ Could not connect to the database. Please check your connection "
            f"settings.\n\nDetails: `{exc.__class__.__name__}: {exc}`"
        )
    except Exception as exc:  # noqa: BLE001 - surface any other startup issue
        st.error(f"❌ Unexpected error while connecting to the database: {exc}")
    return None


# --------------------------------------------------------------------------- #
# Schema introspection helpers (cached so we don't re-hit the DB every rerun)
# --------------------------------------------------------------------------- #
@st.cache_data(ttl=300, show_spinner=False)
def list_tables(_engine: Engine) -> list[str]:
    """List all tables in the 'public' schema."""
    inspector = inspect(_engine)
    return sorted(inspector.get_table_names(schema="public"))


@st.cache_data(ttl=300, show_spinner=False)
def get_columns(_engine: Engine, table_name: str) -> list[dict[str, Any]]:
    """Return column metadata (name + SQLAlchemy/py type) for a table."""
    inspector = inspect(_engine)
    return inspector.get_columns(table_name, schema="public")


def classify_column(col_type: Any) -> str:
    """
    Classify a SQLAlchemy column type into one of: 'date', 'numeric', 'text'.
    Falls back to 'text' for anything unrecognized (safe default).
    """
    type_str = str(col_type).upper()
    date_markers = ("DATE", "TIME", "TIMESTAMP")
    numeric_markers = (
        "INT",
        "NUMERIC",
        "DECIMAL",
        "FLOAT",
        "DOUBLE",
        "REAL",
        "MONEY",
        "SERIAL", 
    )
    if any(marker in type_str for marker in date_markers):
        return "date"
    if any(marker in type_str for marker in numeric_markers):
        return "numeric"
    return "text"


@st.cache_data(ttl=300, show_spinner=False)
def get_distinct_values(_engine: Engine, table_name: str, column_name: str, limit: int = 1000) -> list[Any]:
    """Fetch distinct non-null values for a categorical column (capped)."""
    query = text(
        f'SELECT DISTINCT "{column_name}" FROM "public"."{table_name}" '
        f'WHERE "{column_name}" IS NOT NULL ORDER BY "{column_name}" LIMIT :limit'
    )
    with _engine.connect() as conn:
        result = conn.execute(query, {"limit": limit})
        return [row[0] for row in result]


@st.cache_data(ttl=300, show_spinner=False)
def get_available_years(_engine: Engine, table_name: str, column_name: str) -> list[int]:
    """Fetch distinct years present in a date/timestamp column."""
    query = text(
        f'SELECT DISTINCT EXTRACT(YEAR FROM "{column_name}")::int AS yr '
        f'FROM "public"."{table_name}" WHERE "{column_name}" IS NOT NULL ORDER BY yr'
    )
    with _engine.connect() as conn:
        result = conn.execute(query)
        return [row[0] for row in result]


@st.cache_data(ttl=300, show_spinner=False)
def get_available_year_months(_engine: Engine, table_name: str, column_name: str) -> list[tuple[int, int]]:
    """Fetch distinct (year, month) pairs present in a date/timestamp column."""
    query = text(
        f'SELECT DISTINCT EXTRACT(YEAR FROM "{column_name}")::int AS yr, '
        f'EXTRACT(MONTH FROM "{column_name}")::int AS mo '
        f'FROM "public"."{table_name}" WHERE "{column_name}" IS NOT NULL '
        f'ORDER BY yr, mo'
    )
    with _engine.connect() as conn:
        result = conn.execute(query)
        return [(row[0], row[1]) for row in result]


MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
_MONTH_FULL_TO_NUM = {name: i + 1 for i, name in enumerate(MONTH_NAMES)}
_MONTH_ABBR_TO_NUM = {name[:3]: i + 1 for i, name in enumerate(MONTH_NAMES)}

# --------------------------------------------------------------------------- #
# Crore formatting helper
# --------------------------------------------------------------------------- #
CRORE = 10_000_000  # 1,00,00,000


def format_amount(value: float, in_crores: bool, decimals: int = 2) -> str:
    """Format a numeric total either as-is or converted to Crores (÷1,00,00,000)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "0.00 Cr" if in_crores else "0.00"
    if in_crores:
        return f"{value / CRORE:,.{decimals}f} Cr"
    return f"{value:,.{decimals}f}"


def series_to_crores(series: pd.Series) -> pd.Series:
    """Convert a numeric pandas Series to Crores."""
    return series / CRORE


@st.cache_data(ttl=300, show_spinner=False)
def get_monthly_summary(
    _engine: Engine,
    table_name: str,
    date_col: str,
    category_col: str | None = None,
    sum_cols: tuple[str, ...] = (),
) -> pd.DataFrame:
    """
    Aggregate row counts (and, optionally, SUM of one or more numeric columns)
    per calendar month based on `date_col`, optionally broken down by a second
    column (e.g. payment_mode). A record dated 01-03-2024 is grouped under
    March 2024 regardless of the day of month — this mirrors
    EXTRACT(MONTH FROM ...), same as the existing Month/Year filter above.

    Each requested sum column `col` shows up in the result as `sum_<col>`.
    """
    select_parts = [
        f'EXTRACT(YEAR FROM "{date_col}")::int AS yr',
        f'EXTRACT(MONTH FROM "{date_col}")::int AS mo',
    ]
    group_parts = ["yr", "mo"]

    if category_col:
        select_parts.append(f'"{category_col}"')
        group_parts.append(f'"{category_col}"')

    select_parts.append("COUNT(*) AS record_count")
    for col in sum_cols:
        select_parts.append(f'SUM("{col}") AS "sum_{col}"')

    query = text(
        f'SELECT {", ".join(select_parts)} '
        f'FROM "public"."{table_name}" '
        f'WHERE "{date_col}" IS NOT NULL '
        f'GROUP BY {", ".join(group_parts)} '
        f'ORDER BY yr, mo'
    )
    with _engine.connect() as conn:
        return pd.read_sql(query, conn)


def render_labeled_bar_chart(
    df: pd.DataFrame,
    x_col: str,
    y_col: str,
    color_col: str | None,
    title: str,
    filename_prefix: str,
    key_suffix: str,
    res_scale: int = 3,
    y_label: str = "Count",
    value_format: str = ",",
) -> None:
    """
    Render a Plotly bar chart with the value of each bar printed on top of
    it, styled for high-resolution export. The base canvas is 1600x800px;
    at res_scale=3 the PNG export is 4800x2400 and at res_scale=4 it's
    6400x3200 — both comfortably exceed 4K (3840x2160).

    `y_label` controls the y-axis/legend title text (e.g. "Sum of emi (Cr)")
    and `value_format` controls the on-bar number format (Plotly d3-format
    spec) -- use ",.2f" for decimal amounts like Crore totals, or the
    default "," for whole-number counts.
    """
    fig = px.bar(
        df,
        x=x_col,
        y=y_col,
        color=color_col,
        barmode="group" if color_col else "relative",
        template="plotly_white",
        text=y_col,
        labels={x_col: x_col.replace("_", " ").title(), y_col: y_label},
        title=title,
    )
    fig.update_traces(
        texttemplate="%{text:" + value_format + "}",
        textposition="outside",
        textfont_size=13,
        cliponaxis=False,
    )
    fig.update_layout(
        width=1600,
        height=800,
        font=dict(size=14),
        title_font_size=18,
        uniformtext_minsize=10,
        uniformtext_mode="hide",
        margin=dict(t=80, l=60, r=40, b=60),
        legend_title_text=color_col or "",
        yaxis_title=y_label,
    )

    config = {
        "displaylogo": False,
        "toImageButtonOptions": {
            "format": "png",
            "filename": filename_prefix,
            "scale": res_scale,
        },
    }
    st.plotly_chart(fig, use_container_width=True, config=config, key=f"chart_{key_suffix}")
    st.caption(
        "💡 Click the camera icon in the chart's toolbar (top-right of the chart on hover) "
        f"to export a high-resolution PNG at ≈{1600 * res_scale}×{800 * res_scale}px."
    )


def render_split_bar_charts(
    df: pd.DataFrame,
    x_col: str,
    y_col: str,
    color_col: str | None,
    title: str,
    filename_prefix: str,
    key_suffix: str,
    res_scale: int = 3,
    num_splits: int = 1,
    y_label: str = "Count",
    value_format: str = ",",
) -> None:
    """
    Render one bar chart, or break the x-axis categories (e.g. months) into
    `num_splits` roughly-equal groups and render one chart per group. This
    keeps a chart with many bars from becoming too congested/cluttered to
    read — each smaller chart gets its own title, key, and PNG filename.
    """
    x_order = df[x_col].drop_duplicates().tolist()
    n = len(x_order)

    if num_splits <= 1 or n <= num_splits:
        render_labeled_bar_chart(
            df, x_col, y_col, color_col, title, filename_prefix, key_suffix, res_scale,
            y_label=y_label, value_format=value_format,
        )
        return

    chunk_size = math.ceil(n / num_splits)
    for i in range(0, n, chunk_size):
        chunk_labels = x_order[i : i + chunk_size]
        chunk_df = df[df[x_col].isin(chunk_labels)]
        part_num = i // chunk_size + 1
        render_labeled_bar_chart(
            chunk_df,
            x_col=x_col,
            y_col=y_col,
            color_col=color_col,
            title=f"{title} — Part {part_num} ({chunk_labels[0]} to {chunk_labels[-1]})",
            filename_prefix=f"{filename_prefix}_part{part_num}",
            key_suffix=f"{key_suffix}_part{part_num}",
            res_scale=res_scale,
            y_label=y_label,
            value_format=value_format,
        )


# --------------------------------------------------------------------------- #
# Query builder
# --------------------------------------------------------------------------- #
def build_query(
    table_name: str,
    ref_column: str,
    output_columns: list[str],
    ref_kind: str,
    filter_payload: dict[str, Any],
) -> tuple[str, dict[str, Any]]:
    """
    Build a parameterized SQL query (string with :named placeholders + params
    dict) based on the reference column type and the filters chosen in the UI.
    Always includes the reference column itself in the SELECT list so it's
    visible in the results alongside the requested output columns.
    """
    select_cols = list(dict.fromkeys([ref_column, *output_columns]))  # de-dupe, keep order
    quoted_select = ", ".join(f'"{c}"' for c in select_cols)
    base_query = f'SELECT {quoted_select} FROM "public"."{table_name}"'
    params: dict[str, Any] = {}
    where_clause = ""

    if ref_kind == "date":
        granularity = filter_payload.get("granularity")

        if granularity == "Specific Date(s)":
            dates: list[date] = filter_payload.get("dates", [])
            if dates:
                if len(dates) == 2:
                    # If two dates are selected, use BETWEEN to get the full range
                    params["start_date"] = dates[0]
                    params["end_date"] = dates[1]
                    where_clause = f'WHERE "{ref_column}"::date BETWEEN :start_date AND :end_date'
                elif len(dates) == 1:
                    # If only one date is selected, get that exact day
                    params["single_date"] = dates[0]
                    where_clause = f'WHERE "{ref_column}"::date = :single_date'
                else:
                    # Fallback just in case
                    placeholders = []
                    for i, d in enumerate(dates):
                        key = f"date_{i}"
                        placeholders.append(f":{key}")
                        params[key] = d
                    where_clause = f'WHERE "{ref_column}"::date IN ({", ".join(placeholders)})'

        elif granularity == "Month/Year":
            year_months: list[tuple[int, int]] = filter_payload.get("year_months", [])
            if year_months:
                conditions = []
                for i, (yr, mo) in enumerate(year_months):
                    yr_key, mo_key = f"ym_yr_{i}", f"ym_mo_{i}"
                    conditions.append(
                        f'(EXTRACT(YEAR FROM "{ref_column}") = :{yr_key} '
                        f'AND EXTRACT(MONTH FROM "{ref_column}") = :{mo_key})'
                    )
                    params[yr_key] = yr
                    params[mo_key] = mo
                where_clause = "WHERE " + " OR ".join(conditions)

        elif granularity == "Year(s)":
            years: list[int] = filter_payload.get("years", [])
            if years:
                placeholders = []
                for i, yr in enumerate(years):
                    key = f"year_{i}"
                    placeholders.append(f":{key}")
                    params[key] = yr
                where_clause = f'WHERE EXTRACT(YEAR FROM "{ref_column}") IN ({", ".join(placeholders)})'

    else:  # categorical / text
        values: list[Any] = filter_payload.get("values", [])
        if values:
            placeholders = []
            for i, v in enumerate(values):
                key = f"val_{i}"
                placeholders.append(f":{key}")
                params[key] = v
            where_clause = f'WHERE "{ref_column}" IN ({", ".join(placeholders)})'

    query = f"{base_query} {where_clause}".strip()
    return query, params


@st.cache_data(ttl=60, show_spinner=False)
def run_query(_engine: Engine, query: str, params: dict[str, Any]) -> pd.DataFrame:
    """Execute the (already-parameterized) query and return a DataFrame."""
    with _engine.connect() as conn:
        return pd.read_sql(text(query), conn, params=params)


# --------------------------------------------------------------------------- #
# Excel-style 4-Quadrant Pivot Table engine
# --------------------------------------------------------------------------- #
PIVOT_AGG_FUNCS = ["sum", "mean", "count", "min", "max"]

# Column names used internally for the four pre-aggregated statistics that
# every pivot view (Summary Matrix, Field Efficiency Table, Chart, flat CSV)
# is built from. Carrying all four -- not just whichever agg_func is
# currently selected -- means switching the Values aggregation in the
# sidebar (sum -> mean -> count -> ...) never triggers a new database
# round-trip: it's just a different combination of these same four numbers,
# computed in pandas over an already-tiny, already-cached table.
STAT_SUM = "__stat_sum__"
STAT_COUNT = "__stat_count__"
STAT_MIN = "__stat_min__"
STAT_MAX = "__stat_max__"


@st.cache_data(ttl=120, show_spinner=False)
def fetch_pivot_source_data(
    _engine: Engine,
    table_name: str,
    row_cols: tuple[str, ...],
    col_cols: tuple[str, ...],
    value_col: str,
    filter_conditions: tuple[tuple[str, tuple[Any, ...]], ...],
    column_kind_map: dict[str, str],
) -> pd.DataFrame:
    """
    Push the aggregation into PostgreSQL instead of pulling raw rows.

    One GROUP BY query computes SUM/COUNT/MIN/MAX of `value_col` per
    (Rows x Columns) combination *inside the database* -- including
    auto-bucketing any date/timestamp Rows/Columns field into a Month level
    above the exact Date (mirroring Excel's automatic date grouping), via a
    `TO_CHAR(...)` expression right in the GROUP BY. For realistic BI
    dimensions this turns a 4M+ row table scan into a query that returns at
    most a few hundred/thousand aggregated rows.

    This is the single biggest lever for large tables: every downstream
    view -- Summary Matrix, Field Efficiency Table, Chart, flat CSV --
    is then built from that tiny table with fast vectorized pandas
    groupbys, never touching millions of raw rows in Python again.

    Returns a DataFrame with one row per unique (expanded) Rows x Columns
    combination, containing:
      - one column per Rows/Columns field (date fields become TWO string
        columns, "__<field>__month" ('YYYY-MM') and "__<field>__date"
        ('YYYY-MM-DD') -- exactly the field-name convention `_expand_date_fields`
        already expects, so no downstream code needs to know this was
        computed in SQL rather than in pandas)
      - STAT_SUM / STAT_COUNT / STAT_MIN / STAT_MAX for `value_col`.
    """
    group_exprs: list[str] = []
    group_by_aliases: list[str] = []
    seen: set[str] = set()

    for f in [*row_cols, *col_cols]:
        if f in seen:
            continue
        seen.add(f)
        if column_kind_map.get(f) == "date":
            month_alias, date_alias = f"__{f}__month", f"__{f}__date"
            group_exprs.append(f"TO_CHAR(\"{f}\", 'YYYY-MM') AS \"{month_alias}\"")
            group_exprs.append(f"TO_CHAR(\"{f}\", 'YYYY-MM-DD') AS \"{date_alias}\"")
            group_by_aliases.extend([f'"{month_alias}"', f'"{date_alias}"'])
        else:
            group_exprs.append(f'"{f}" AS "{f}"')
            group_by_aliases.append(f'"{f}"')

    agg_exprs = [
        f'SUM("{value_col}") AS "{STAT_SUM}"',
        f'COUNT("{value_col}") AS "{STAT_COUNT}"',
        f'MIN("{value_col}") AS "{STAT_MIN}"',
        f'MAX("{value_col}") AS "{STAT_MAX}"',
    ]
    select_clause = ", ".join([*group_exprs, *agg_exprs])
    query = f'SELECT {select_clause} FROM "public"."{table_name}"'

    params: dict[str, Any] = {}
    where_parts: list[str] = []
    for i, (col, values) in enumerate(filter_conditions):
        if not values:
            continue
        placeholders = []
        for j, v in enumerate(values):
            key = f"pf_{i}_{j}"
            placeholders.append(f":{key}")
            params[key] = v
        where_parts.append(f'"{col}" IN ({", ".join(placeholders)})')
    if where_parts:
        query += " WHERE " + " AND ".join(where_parts)

    if group_by_aliases:
        query += " GROUP BY " + ", ".join(group_by_aliases)

    with _engine.connect() as conn:
        agg_df = pd.read_sql(text(query), conn, params=params)

    for c in (STAT_SUM, STAT_COUNT, STAT_MIN, STAT_MAX):
        agg_df[c] = pd.to_numeric(agg_df[c], errors="coerce")
    return agg_df


def _combine_group_stats(df: pd.DataFrame, group_cols: list[str], agg_func: str) -> pd.Series:
    """
    Combine the pre-aggregated STAT_SUM/STAT_COUNT/STAT_MIN/STAT_MAX columns
    across `group_cols` -- a grouping that may be *coarser* than the leaf
    granularity the stats were computed at, i.e. a subtotal or a margin --
    into ONE correctly-combined value per group for `agg_func`.

    sum/count/min/max all combine correctly by re-applying the SAME
    operation across the leaf-level stats (sum of sums, sum of counts, min
    of mins, max of maxes). mean is reconstructed as sum/count -- NEVER as
    an average of the leaf-level means, which silently gives the wrong
    number the instant the underlying group sizes differ (the classic
    "average of ratios" vs "ratio of sums" trap). `group_cols=[]` combines
    across the ENTIRE table into a single-row Series (used for grand
    totals / the corner cell).
    """
    if group_cols:
        g = df.groupby(group_cols, dropna=False)
    else:
        g = df.groupby(np.zeros(len(df), dtype=int))

    if agg_func == "sum":
        return g[STAT_SUM].sum().astype(float)
    if agg_func == "count":
        return g[STAT_COUNT].sum().astype(float)
    if agg_func == "mean":
        sum_s = g[STAT_SUM].sum()
        count_s = g[STAT_COUNT].sum()
        return (sum_s / count_s.replace(0, np.nan)).fillna(0.0).astype(float)
    if agg_func == "min":
        return g[STAT_MIN].min().astype(float)
    if agg_func == "max":
        return g[STAT_MAX].max().astype(float)
    raise ValueError(f"Unsupported agg_func: {agg_func}")


def _expand_date_fields(
    df: pd.DataFrame,
    fields: list[str],
    column_kind_map: dict[str, str],
    label_map: dict[tuple[str, Any], str],
) -> tuple[pd.DataFrame, list[str]]:
    """
    Mirrors Excel's automatic date grouping: a date/timestamp Rows/Columns
    field expands into two levels -- Month ('Mon-YY') above the exact Date
    ('DD-Mon-YY'). The bucketed columns themselves ("__<field>__month" /
    "__<field>__date", both plain 'YYYY-MM' / 'YYYY-MM-DD' text) are now
    computed server-side by `fetch_pivot_source_data`'s GROUP BY -- this
    function's job is just to know the resulting effective field names and
    build the pretty display `label_map` from those already-bucketed text
    columns. Non-date fields pass through unchanged. `df` is returned
    as-is (no mutation needed any more); kept in the signature/return so
    every existing call site keeps working unchanged.
    """
    expanded: list[str] = []
    for f in fields:
        if column_kind_map.get(f) == "date":
            month_field, date_field = f"__{f}__month", f"__{f}__date"
            if month_field in df.columns:
                month_keys = df[month_field].dropna().unique()
                if len(month_keys):
                    disp = pd.to_datetime(month_keys, format="%Y-%m").strftime("%b-%y")
                    for k, d in zip(month_keys, disp):
                        label_map[(month_field, k)] = d
            if date_field in df.columns:
                date_keys = df[date_field].dropna().unique()
                if len(date_keys):
                    disp = pd.to_datetime(date_keys, format="%Y-%m-%d").strftime("%d-%b-%y")
                    for k, d in zip(date_keys, disp):
                        label_map[(date_field, k)] = d
            expanded.extend([month_field, date_field])
        else:
            expanded.append(f)
    return df, expanded


def _pivot_agg_value(
    df: pd.DataFrame,
    agg_func: str,
    row_fields: list[str],
    row_prefix: tuple[Any, ...],
    col_fields: list[str],
    col_prefix: tuple[Any, ...],
) -> float:
    """Aggregate over exactly the (already tiny, pre-aggregated) rows matching
    row_prefix + col_prefix -- a *partial* prefix aggregates over every deeper
    level, which is what makes subtotal cells correct for sum/mean/count/
    min/max alike. Operates on STAT_SUM/STAT_COUNT/STAT_MIN/STAT_MAX, never
    on raw per-record data, so this stays fast even called thousands of
    times while building a large pivot."""
    mask = pd.Series(True, index=df.index)
    for f, v in zip(row_fields, row_prefix):
        mask &= df[f] == v
    for f, v in zip(col_fields, col_prefix):
        mask &= df[f] == v
    subset = df.loc[mask]
    if subset.empty:
        return 0.0
    if agg_func == "sum":
        return float(subset[STAT_SUM].sum())
    if agg_func == "count":
        return float(subset[STAT_COUNT].sum())
    if agg_func == "mean":
        total_count = subset[STAT_COUNT].sum()
        return float(subset[STAT_SUM].sum() / total_count) if total_count else 0.0
    if agg_func == "min":
        return float(subset[STAT_MIN].min())
    if agg_func == "max":
        return float(subset[STAT_MAX].max())
    raise ValueError(f"Unsupported agg_func: {agg_func}")


def _month_sort_key(value: Any) -> tuple[int, int, str] | None:
    """
    If `value` looks like a month name someone typed/stored as plain text —
    'January', 'Jan', 'January 2024', 'Jan-24', 'Jan 24' — return a
    (year, month, text) key so it sorts chronologically (Jan -> Dec, and by
    year first if a year is present) instead of alphabetically. Returns None
    for anything that doesn't look like a month value, so normal columns are
    completely unaffected.
    """
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None
    match = re.match(r"^([A-Za-z]+)[\s\-,]*(\d{2,4})?$", text)
    if not match:
        return None
    month_part, year_part = match.group(1), match.group(2)
    month_num = _MONTH_FULL_TO_NUM.get(month_part.title()) or _MONTH_ABBR_TO_NUM.get(month_part[:3].title())
    if month_num is None:
        return None
    if year_part:
        year_num = int(year_part)
        if year_num < 100:  # 2-digit year, e.g. "24" -> 2024
            year_num += 2000
    else:
        year_num = 0  # no year given -> just order Jan..Dec
    return (year_num, month_num, text)


def _ordered_unique(series: pd.Series) -> list[Any]:
    """
    Distinct non-null values, sorted when possible. If EVERY value in the
    column looks like a month name (plain text, not a real date/timestamp
    column — those are already handled separately and sort correctly), sort
    chronologically (January -> December, and by year first if present)
    instead of alphabetically. Otherwise falls back to a normal sort, or
    first-seen order if the values aren't sortable at all.
    """
    vals = [v for v in pd.unique(series) if pd.notna(v)]
    if not vals:
        return vals
    month_keys = [_month_sort_key(v) for v in vals]
    if all(k is not None for k in month_keys):
        return [v for _, v in sorted(zip(month_keys, vals), key=lambda pair: pair[0])]
    try:
        return sorted(vals)
    except TypeError:
        return vals


def _reorder_axis_chronologically(idx: pd.Index) -> pd.Index:
    """
    Reorder a (possibly MultiIndex) pandas axis so any month-like level
    sorts chronologically instead of alphabetically — the exact same rule
    `_ordered_unique` applies to the main Summary Matrix — while pinning a
    literal 'Total' entry at that level to the very end. Non-month levels
    keep their existing order untouched. Used for tables (like the Field
    Efficiency Table) that are built via pandas groupby/pivot_table instead
    of the row/column tree builder, so they still need this applied
    separately after the fact.
    """
    is_multi = isinstance(idx, pd.MultiIndex)
    n_levels = idx.nlevels if is_multi else 1
    tuples = list(idx) if is_multi else [(v,) for v in idx]

    level_ranks: list[dict[Any, int]] = []
    for level in range(n_levels):
        non_total_vals = [t[level] for t in tuples if t[level] not in ("Total", "")]
        ordered_vals = _ordered_unique(pd.Series(non_total_vals)) if non_total_vals else []
        level_ranks.append({v: i for i, v in enumerate(ordered_vals)})

    def _sort_key(t: tuple) -> tuple:
        key = []
        for level in range(n_levels):
            v = t[level]
            key.append((1, 0) if v in ("Total", "") else (0, level_ranks[level].get(v, 0)))
        return tuple(key)

    order = sorted(range(len(tuples)), key=lambda i: _sort_key(tuples[i]))
    new_tuples = [tuples[i] for i in order]

    if is_multi:
        return pd.MultiIndex.from_tuples(new_tuples, names=idx.names)
    return pd.Index([t[0] for t in new_tuples], name=idx.name)


def _build_row_entries(
    df: pd.DataFrame,
    row_fields: list[str],
    label_map: dict[tuple[str, Any], str],
    prefix: tuple[Any, ...] = (),
    indent: int = 0,
) -> list[dict[str, Any]]:
    """
    Excel-style row tree: a subtotal row for each parent group appears
    BEFORE its children (e.g. 'EAST' subtotal, then 'EAST_1', 'EAST_2'
    indented beneath it), recursing to arbitrary depth.
    """
    level = len(prefix)
    if level >= len(row_fields):
        return []
    field = row_fields[level]
    mask = pd.Series(True, index=df.index)
    for f, v in zip(row_fields[:level], prefix):
        mask &= df[f] == v
    values = _ordered_unique(df.loc[mask, field])
    is_last_level = level == len(row_fields) - 1

    entries: list[dict[str, Any]] = []
    for v in values:
        new_prefix = prefix + (v,)
        disp = label_map.get((field, v), str(v))
        if is_last_level:
            entries.append({"label": disp, "prefix": new_prefix, "indent": indent, "is_group": False})
        else:
            entries.append({"label": disp, "prefix": new_prefix, "indent": indent, "is_group": True})
            entries.extend(_build_row_entries(df, row_fields, label_map, new_prefix, indent + 1))
    return entries


def _build_col_entries(
    df: pd.DataFrame,
    col_fields: list[str],
    label_map: dict[tuple[str, Any], str],
    prefix: tuple[Any, ...] = (),
) -> list[dict[str, Any]]:
    """
    Excel-style column tree: a group's children come first, followed by that
    group's own '<value> Total' subtotal column (e.g. Jun, Jun Total, Jul,
    Jul Total, then STAGE_1 Total) — the mirror image of the row ordering.
    """
    level = len(prefix)
    if level >= len(col_fields):
        return []
    field = col_fields[level]
    mask = pd.Series(True, index=df.index)
    for f, v in zip(col_fields[:level], prefix):
        mask &= df[f] == v
    values = _ordered_unique(df.loc[mask, field])
    is_last_level = level == len(col_fields) - 1

    entries: list[dict[str, Any]] = []
    for v in values:
        new_prefix = prefix + (v,)
        if is_last_level:
            disp = label_map.get((field, v), str(v))
            entries.append({"label": disp, "prefix": new_prefix, "is_total": False})
        else:
            entries.extend(_build_col_entries(df, col_fields, label_map, new_prefix))
            disp = label_map.get((field, v), str(v))
            entries.append({"label": f"{disp} Total", "prefix": new_prefix, "is_total": True})
    return entries


def _pretty_prefix(
    prefix: tuple[Any, ...], fields: list[str], label_map: dict[tuple[str, Any], str]
) -> tuple[str, ...]:
    """Map each raw value in a prefix tuple to its display label, one per field."""
    return tuple(label_map.get((fields[i], v), str(v)) for i, v in enumerate(prefix))


@st.cache_data(ttl=300, show_spinner=False)
def build_excel_style_pivot(
    source_df: pd.DataFrame,
    row_cols: list[str],
    col_cols: list[str],
    value_col: str,
    agg_func: str,
    column_kind_map: dict[str, str],
    in_crores: bool,
    right_total_mode: str = "grand_total",
) -> tuple[pd.DataFrame, list[bool], list[bool]]:
    """
    Build a true Excel-style PivotTable: per-level row AND column subtotals
    (not just one grand-total margin), date fields auto-grouped into a
    Month level above the exact date, and values optionally shown in
    Crores. Returns (display_df, group_row_flags, total_col_flags) — the
    two flag lists are positional (aligned to display_df's row/column
    order), used only for highlighting subtotal/Total rows & columns.

    `right_total_mode` controls what appears at the far right of the header
    when 2+ Columns fields are selected (e.g. Stage > Month):
      - "grand_total"        -> single overall "Grand Total" column (default,
                                 original behaviour).
      - "deepest_field_total" -> one "<value> Total" column per distinct value
                                 of the *last* Columns field (e.g. one per
                                 month), each summed across every other
                                 Columns field (e.g. across all stages) --
                                 no single overall Grand Total column.
      - "both"                -> the per-deepest-field totals, followed by
                                 one overall Grand Total at the very end.
    With fewer than 2 Columns fields there's nothing to "sum across", so
    this always falls back to the plain single Grand Total column.
    """
    label_map: dict[tuple[str, Any], str] = {}
    work_df, row_fields = _expand_date_fields(source_df, row_cols, column_kind_map, label_map)
    work_df, col_fields = _expand_date_fields(work_df, col_cols, column_kind_map, label_map)

    row_entries = _build_row_entries(work_df, row_fields, label_map)
    row_entries.append({"label": "Grand Total", "prefix": (), "indent": 0, "is_group": True, "is_grand_total": True})

    if col_fields:
        col_entries = _build_col_entries(work_df, col_fields, label_map)
        col_depth = len(col_fields)

        # Dynamic "per-deepest-field" totals (e.g. one "Jul-26 Total" column
        # per month, summed across every Stage) -- only meaningful with 2+
        # Columns fields; with just one field there's no "other" field left
        # to sum across, so this list simply stays empty in that case.
        deepest_field_entries: list[dict[str, Any]] = []
        if right_total_mode in ("deepest_field_total", "both") and col_depth >= 2:
            deepest_field = col_fields[-1]
            for v in _ordered_unique(work_df[deepest_field]):
                disp = label_map.get((deepest_field, v), str(v))
                deepest_field_entries.append({
                    "label": f"{disp} Total",
                    "prefix": (v,),
                    "is_total": True,
                    # Aggregate by matching ONLY this field -- ignoring every
                    # other Columns field entirely, which is what makes this
                    # a total *across* stages rather than a per-stage subtotal.
                    "mask_fields": [deepest_field],
                })

        # Fall back to (or additionally include) the classic single overall
        # Grand Total column -- always included if the per-field totals
        # above ended up empty for any reason (e.g. mode requested but only
        # 1 Columns field was actually picked).
        grand_total_entries: list[dict[str, Any]] = []
        if right_total_mode in ("grand_total", "both") or not deepest_field_entries:
            grand_total_entries = [{"label": "Grand Total", "prefix": (), "is_total": True, "is_grand_total": True}]

        col_entries = col_entries + deepest_field_entries + grand_total_entries

        col_tuples: list[tuple[str, ...]] = []
        for ce in col_entries:
            if ce.get("is_grand_total"):
                tup = ("Grand Total",) + ("",) * (col_depth - 1)
            elif ce.get("mask_fields"):
                # Per-deepest-field total (e.g. "Jul-26 Total") -- shown as a
                # single top-level header label, same visual treatment as
                # Grand Total, since it cuts across every other Columns
                # field rather than nesting under one specific value of them.
                tup = (ce["label"],) + ("",) * (col_depth - 1)
            else:
                depth_used = len(ce["prefix"])
                pretty = _pretty_prefix(ce["prefix"], col_fields[:depth_used], label_map)
                if ce["is_total"]:
                    pretty = pretty[:-1] + (f"{pretty[-1]} Total",)
                tup = pretty + ("",) * (col_depth - len(pretty))
            col_tuples.append(tup)
        columns_index = pd.MultiIndex.from_tuples(col_tuples) if col_depth > 1 else pd.Index([t[0] for t in col_tuples])
        total_col_flags = [bool(ce.get("is_total") or ce.get("is_grand_total")) for ce in col_entries]
    else:
        col_entries = [{"label": f"{agg_func.title()} of {value_col}", "prefix": (), "is_total": False}]
        columns_index = pd.Index([col_entries[0]["label"]])
        total_col_flags = [False]

    divisor = CRORE if (in_crores and agg_func != "count") else 1

    row_labels: list[str] = []
    group_row_flags: list[bool] = []
    data_rows: list[list[float]] = []
    for re_ in row_entries:
        indent_txt = "    " * re_["indent"]
        if re_.get("is_grand_total"):
            label = "Grand Total"
        elif re_["is_group"]:
            label = f"{indent_txt}{re_['label']}"
        else:
            # Only show the child-arrow when this row is actually nested under a
            # parent group row (indent > 0). A single-level Rows selection has
            # no parent, so it should read as a plain flat list, e.g. "EAST",
            # not "↳ EAST".
            arrow = "↳ " if re_["indent"] > 0 else ""
            label = f"{indent_txt}{arrow}{re_['label']}"
        row_labels.append(label)
        group_row_flags.append(bool(re_["is_group"]))

        row_vals = [
            _pivot_agg_value(
                work_df, agg_func, row_fields, re_["prefix"],
                ce.get("mask_fields", col_fields), ce["prefix"],
            ) / divisor
            for ce in col_entries
        ]
        data_rows.append(row_vals)

    display_df = pd.DataFrame(
        data_rows,
        index=pd.Index(row_labels, name=" / ".join(row_cols)),
        columns=columns_index,
    )
    return display_df, group_row_flags, total_col_flags


def _render_styled_pivot_matrix(
    display_df: pd.DataFrame,
    group_row_flags: list[bool],
    total_col_flags: list[bool],
    agg_func: str,
) -> None:
    """Shared styled st.dataframe rendering for any Excel-style Summary
    Matrix (single- or multi-Values-column) -- bolds/shades subtotal rows
    and Total columns, formats numbers, falls back to a plain unstyled
    table if styling fails for any reason rather than crashing the page."""
    try:
        num_fmt = "{:,.0f}" if agg_func == "count" else "{:,.2f}"

        def _apply_styles(df: pd.DataFrame) -> pd.DataFrame:
            styles = pd.DataFrame("", index=df.index, columns=df.columns)
            row_labels = list(df.index)
            for i, (label, is_group) in enumerate(zip(row_labels, group_row_flags)):
                if label == "Grand Total":
                    styles.iloc[i, :] = "font-weight:bold"
                elif is_group:
                    styles.iloc[i, :] = "background-color:#dbe5f1; font-weight:bold"
            for j, is_total in enumerate(total_col_flags):
                if is_total:
                    for i in range(len(df)):
                        existing = styles.iloc[i, j]
                        styles.iloc[i, j] = f"{existing}; background-color:#f2f2f2; font-weight:bold" if existing else "background-color:#f2f2f2; font-weight:bold"
            return styles

        styler = display_df.style.apply(_apply_styles, axis=None).format(num_fmt)
        st.dataframe(styler, use_container_width=True)
    except Exception:
        # If styling fails for any reason (older pandas/streamlit version, etc.),
        # fall back to a plain, still-correct, unstyled table rather than crashing.
        st.dataframe(display_df, use_container_width=True)


_XL_HEADER_FILL = "1F3864"       # formal dark navy -- matches a corporate report header band
_XL_HEADER_FONT = "FFFFFF"
_XL_GROUP_ROW_FILL = "DCE6F1"    # light blue -- matches the on-screen subtotal-row shading
_XL_TOTAL_COL_FILL = "F2F2F2"    # light grey -- matches the on-screen Total-column shading
_XL_TOTAL_ROW_FILL = "DCE6F1"    # Field-Efficiency Total row shading (same family as group rows)
_XL_CORNER_FILL = "B4C6E7"       # Total-row x Total-col intersection, a shade deeper


def dataframe_to_formatted_excel_bytes(
    df: pd.DataFrame,
    group_row_flags: list[bool] | None = None,
    total_col_flags: list[bool] | None = None,
    value_kind: str = "number",  # "number" | "count" | "percent"
    table_style: str = "matrix",  # "matrix" (Summary Matrix rules) | "efficiency" (Field Efficiency rules)
    sheet_name: str = "Report",
) -> bytes:
    """
    Export a (possibly MultiIndex-column) pivot/summary DataFrame to a
    properly formatted .xlsx -- unlike a flattened CSV, an outer header like
    "Sum of March 23" is written as a real merged cell spanning BOUNCED /
    CLEARED beneath it, instead of silently disappearing. Subtotal/Total
    rows and columns get the same bold/shading treatment shown on screen,
    and percent tables (Field Efficiency) are written exactly as displayed,
    e.g. "18.80%", via a custom number format -- not divided by 100 again.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Report")[:31]

    has_multi_cols = isinstance(df.columns, pd.MultiIndex)
    col_levels = df.columns.nlevels if has_multi_cols else 1
    index_is_multi = isinstance(df.index, pd.MultiIndex)
    n_index_cols = df.index.nlevels if index_is_multi else 1
    raw_names = list(df.index.names) if index_is_multi else [df.index.name]
    index_names = [n if n else f"Row {i + 1}" for i, n in enumerate(raw_names)]

    header_font = Font(bold=True, color=_XL_HEADER_FONT, size=11, name="Calibri")
    header_fill = PatternFill("solid", fgColor=_XL_HEADER_FILL)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_cell(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # --- Index (row-key) header, spanning every header row --- #
    for i, name in enumerate(index_names):
        c = ws.cell(row=1, column=1 + i, value=name)
        style_header_cell(c)
        if col_levels > 1:
            ws.merge_cells(start_row=1, start_column=1 + i, end_row=col_levels, end_column=1 + i)
            for r in range(2, col_levels + 1):
                style_header_cell(ws.cell(row=r, column=1 + i))

    data_start_col = 1 + n_index_cols

    # --- Value column headers, merged per level so parent labels (e.g.
    # "Sum of March 23") correctly span every child column beneath them --- #
    col_tuples = list(df.columns) if has_multi_cols else [(c,) for c in df.columns]
    for level in range(col_levels):
        n = len(col_tuples)
        i = 0
        while i < n:
            j = i
            key = col_tuples[i][: level + 1]
            while j + 1 < n and col_tuples[j + 1][: level + 1] == key:
                j += 1
            label = col_tuples[i][level]
            col0 = data_start_col + i
            cell = ws.cell(row=level + 1, column=col0, value=("" if label is None else label))
            style_header_cell(cell)
            if j > i:
                ws.merge_cells(start_row=level + 1, start_column=col0, end_row=level + 1, end_column=data_start_col + j)
                for extra_col in range(col0 + 1, data_start_col + j + 1):
                    style_header_cell(ws.cell(row=level + 1, column=extra_col))
            i = j + 1

    # --- Body --- #
    num_fmt = {'percent': '0.00"%"', 'count': '#,##0'}.get(value_kind, '#,##0.00')
    body_font = Font(name="Calibri", size=10.5)
    group_fill = PatternFill("solid", fgColor=_XL_GROUP_ROW_FILL)
    total_col_fill = PatternFill("solid", fgColor=_XL_TOTAL_COL_FILL)
    total_row_fill = PatternFill("solid", fgColor=_XL_TOTAL_ROW_FILL)
    corner_fill = PatternFill("solid", fgColor=_XL_CORNER_FILL)

    flat_index_tuples = list(df.index) if index_is_multi else [(v,) for v in df.index]
    n_rows = len(df)
    header_rows = col_levels

    def is_grand_total_row(r: int) -> bool:
        label = str(flat_index_tuples[r][-1]).strip()
        return label in ("Grand Total", "Total")

    for r in range(n_rows):
        excel_row = header_rows + 1 + r
        is_group = bool(group_row_flags[r]) if group_row_flags is not None and r < len(group_row_flags) else False
        is_total_row = is_grand_total_row(r) or (table_style == "efficiency" and r == n_rows - 1)

        # -- index cells -- #
        for c in range(n_index_cols):
            val = flat_index_tuples[r][c] if c < len(flat_index_tuples[r]) else ""
            cell = ws.cell(row=excel_row, column=1 + c, value=val)
            cell.border = border
            if table_style == "matrix":
                cell.font = Font(name="Calibri", size=10.5, bold=(is_total_row or is_group))
                if is_group and not is_total_row:
                    cell.fill = group_fill
            else:  # efficiency table
                cell.font = Font(name="Calibri", size=10.5, bold=is_total_row)
                if is_total_row:
                    cell.fill = total_row_fill

        # -- data cells -- #
        for j in range(df.shape[1]):
            val = df.iat[r, j]
            is_total_col = bool(total_col_flags[j]) if total_col_flags is not None and j < len(total_col_flags) else False
            cell = ws.cell(row=excel_row, column=data_start_col + j, value=(None if pd.isna(val) else float(val)))
            cell.number_format = num_fmt
            cell.border = border

            if table_style == "matrix":
                # Matches the on-screen rule: Grand Total row = bold only (no
                # fill); subtotal/group rows = blue fill + bold; Total
                # columns = grey fill + bold, layered on top of either.
                cell.font = Font(name="Calibri", size=10.5, bold=(is_total_row or is_group or is_total_col))
                if is_group and not is_total_row:
                    cell.fill = group_fill
                if is_total_col:
                    cell.fill = total_col_fill
            else:  # efficiency table: Total row/col shaded, corner a shade deeper
                cell.font = Font(name="Calibri", size=10.5, bold=(is_total_row or is_total_col))
                if is_total_row and is_total_col:
                    cell.fill = corner_fill
                elif is_total_row:
                    cell.fill = total_row_fill
                elif is_total_col:
                    cell.fill = total_col_fill

    ws.freeze_panes = ws.cell(row=header_rows + 1, column=data_start_col)

    # -- column widths -- #
    for c in range(n_index_cols):
        longest = max([len(str(index_names[c]))] + [len(str(t[c])) for t in flat_index_tuples if c < len(t)])
        ws.column_dimensions[get_column_letter(1 + c)].width = min(38, max(12, longest + 2))
    for j in range(df.shape[1]):
        header_len = max(len(str(x)) for x in col_tuples[j]) if col_tuples[j] else 8
        ws.column_dimensions[get_column_letter(data_start_col + j)].width = min(22, max(11, header_len + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_excel_style_pivot_table(
    source_df: pd.DataFrame,
    row_cols: list[str],
    col_cols: list[str],
    value_col: str,
    agg_func: str,
    column_kind_map: dict[str, str],
    in_crores: bool,
    right_total_mode: str = "grand_total",
) -> pd.DataFrame:
    """Build the Excel-style nested pivot, render it (styled where possible), return it for CSV export."""
    display_df, group_row_flags, total_col_flags = build_excel_style_pivot(
        source_df, row_cols, col_cols, value_col, agg_func, column_kind_map, in_crores, right_total_mode
    )

    crore_suffix = " (in Cr)" if (in_crores and agg_func != "count") else ""
    st.markdown(f"**{agg_func.title()} of {value_col}{crore_suffix}**")
    _render_styled_pivot_matrix(display_df, group_row_flags, total_col_flags, agg_func)
    return display_df


def render_multi_value_pivot_table(
    source_df_by_value_col: dict[str, pd.DataFrame],
    row_cols: list[str],
    col_cols: list[str],
    value_cols: list[str],
    agg_func: str,
    column_kind_map: dict[str, str],
    in_crores: bool,
    right_total_mode: str = "grand_total",
) -> pd.DataFrame:
    """
    Excel's "multiple Values fields" behaviour: build the IDENTICAL
    Excel-style pivot (same Rows/Columns/right_total_mode -- only the
    underlying numbers differ) once per Values column, using the exact same
    tested single-value engine, then place the results side by side with an
    extra outer header level naming each one (e.g. "Sum of EMI_Mar23" |
    "Sum of EMI_Apr23") -- exactly what Excel shows when you drag more than
    one field into the Values area. Every block is explicitly reindexed to
    the FIRST block's row order before concatenating (guards against the
    rare case where two separate SQL queries return their groups in a
    slightly different order), so this is a safe, purely side-by-side
    placement -- never an averaged, re-derived, or misaligned number.
    """
    blocks: dict[str, pd.DataFrame] = {}
    group_row_flags: list[bool] = []
    tiled_total_col_flags: list[bool] = []
    base_index: pd.Index | None = None

    for vc in value_cols:
        block_df, block_group_row_flags, block_total_col_flags = build_excel_style_pivot(
            source_df_by_value_col[vc], row_cols, col_cols, vc, agg_func, column_kind_map, in_crores, right_total_mode
        )
        if base_index is None:
            base_index = block_df.index
            group_row_flags = block_group_row_flags
        else:
            block_df = block_df.reindex(base_index).fillna(0.0)
        blocks[f"{agg_func.title()} of {vc}"] = block_df
        tiled_total_col_flags.extend(block_total_col_flags)

    display_df = pd.concat(blocks, axis=1)

    crore_suffix = " (in Cr)" if (in_crores and agg_func != "count") else ""
    st.markdown(
        f"**{agg_func.title()} of {len(value_cols)} Values columns{crore_suffix}** "
        "— one column block per field, side by side"
    )
    _render_styled_pivot_matrix(display_df, group_row_flags, tiled_total_col_flags, agg_func)
    return display_df


def _effective_col_level(col_cols: list[str], target_field: str, column_kind_map: dict[str, str]) -> int:
    """
    Map a field name in `col_cols` to its column-level index in the *effective*
    (date-expansion-aware) column tree built by `build_excel_style_pivot`. A
    date/timestamp field expands into TWO levels (Month, then exact Date) via
    `_expand_date_fields`, so any field listed after one in `col_cols` is
    shifted by an extra level for each date field preceding it. Non-date
    fields (like a DPD slab column) occupy exactly one level.
    """
    level = 0
    for f in col_cols:
        if f == target_field:
            return level
        level += 2 if column_kind_map.get(f) == "date" else 1
    raise ValueError(f"'{target_field}' is not one of the selected Columns fields: {col_cols}")


def add_dpd_buckets_to_excel_pivot(
    display_df: pd.DataFrame, slab_level: int, only_show_buckets: bool = False
) -> pd.DataFrame:
    """
    Append '1+', '30+', '90+' cumulative DPD macro-bucket column groups to a
    display_df produced by `build_excel_style_pivot` / `render_excel_style_pivot_table`
    (the Excel-style Summary Matrix) -- the table that already contains
    per-slab subtotal columns like '1-29 Total' and a 'Grand Total' column.

    '1+'  = sum of '1-29' + '30-59' + '60-89' + '90+'
    '30+' = sum of '30-59' + '60-89' + '90+'
    '90+' = sum of '90+' alone -- when the raw slabs are being kept
            (only_show_buckets=False) this is IDENTICAL to the existing raw
            '90+' slab column, so it's skipped rather than added as a
            confusing duplicate column with the same label. When the raw
            slabs are being dropped (only_show_buckets=True) there's no
            collision any more, so '90+' is added normally.

    `slab_level` is the column level (from `_effective_col_level`) holding the
    raw DPD slab values. When 2+ Columns fields are selected, the DPD slab
    field currently must be the FIRST Columns field (slab_level == 0) so the
    new '<bucket> Total' columns line up with the existing '<slab> Total'
    subtotal convention; with just one Columns field (flat columns) any
    position works since there's nothing else to reconstruct.

    only_show_buckets : bool, default False
        If True, the raw slab columns ('1-29', '30-59', '60-89', '90+') and
        their '<slab> Total' subtotals are removed from the output entirely,
        leaving only '1+' / '30+' / '90+' (plus 'Grand Total' and any other
        unrelated columns untouched).

    Raises ValueError (caught and shown as a friendly warning by the caller)
    if the shape isn't supported, or if no raw DPD-slab columns are found.
    """
    columns = display_df.columns
    is_multi = isinstance(columns, pd.MultiIndex)
    n_levels = columns.nlevels if is_multi else 1

    if n_levels > 1 and slab_level != 0:
        raise ValueError(
            "Adding DPD buckets currently requires the DPD slab field to be the "
            "FIRST field in your Columns list when 2+ Columns fields are selected. "
            "Reorder your Columns selection in the sidebar so the DPD slab field "
            "comes first, then try again."
        )

    known_slabs = {"1-29", "30-59", "60-89", "90+"}
    bucket_defs = {
        "1+": {"1-29", "30-59", "60-89", "90+"},
        "30+": {"30-59", "60-89", "90+"},
        "90+": {"90+"},
    }

    tuples = [t if isinstance(t, tuple) else (t,) for t in columns]

    def _slab_val(t: tuple) -> Any:
        return t[slab_level] if is_multi else t[0]

    def _is_grand_total(t: tuple) -> bool:
        return any(str(v) == "Grand Total" for v in t)

    def _is_group_total(t: tuple) -> bool:
        return any(isinstance(v, str) and v.endswith(" Total") and v != "Grand Total" for v in t)

    def _is_slab_related(t: tuple) -> bool:
        """True for a raw leaf slab column OR its '<slab> Total' subtotal."""
        v = str(_slab_val(t))
        if v in known_slabs:
            return True
        if v.endswith(" Total") and v[: -len(" Total")] in known_slabs:
            return True
        return False

    leaf_tuples = [
        t for t in tuples
        if not _is_grand_total(t) and not _is_group_total(t) and str(_slab_val(t)) in known_slabs
    ]
    if not leaf_tuples:
        raise ValueError(
            "No raw DPD-slab columns ('1-29', '30-59', '60-89', '90+') were found in the "
            "selected field -- double-check you picked the right Columns field."
        )

    existing_leaf_slabs = {str(_slab_val(t)) for t in leaf_tuples}

    def _other_key(t: tuple) -> tuple:
        return tuple(v for i, v in enumerate(t) if i != slab_level)

    seen_other: list[tuple] = []
    seen_set: set = set()
    for t in leaf_tuples:
        ok = _other_key(t)
        if ok not in seen_set:
            seen_set.add(ok)
            seen_other.append(ok)

    new_frames = []
    for bucket_name, slabs in bucket_defs.items():
        # Skip a bucket that's identical to an existing raw slab column --
        # but only when that raw slab column is still going to be shown.
        if not only_show_buckets and bucket_name in existing_leaf_slabs and slabs == {bucket_name}:
            continue

        bucket_leaf = [t for t in leaf_tuples if str(_slab_val(t)) in slabs]
        if not bucket_leaf:
            continue

        block_cols: list[Any] = []
        block_data: dict[Any, pd.Series] = {}
        for ok in seen_other:
            matching = [t for t in bucket_leaf if _other_key(t) == ok]
            if not matching:
                continue
            summed = display_df[matching].sum(axis=1)
            template = list(matching[0])
            template[slab_level] = bucket_name
            new_tuple = tuple(template) if is_multi else template[0]
            block_data[new_tuple] = summed
            block_cols.append(new_tuple)

        block_df = pd.DataFrame(block_data)[block_cols]

        if n_levels > 1:
            total_tuple = (f"{bucket_name} Total",) + ("",) * (n_levels - 1)
            block_df[total_tuple] = block_df.sum(axis=1)

        new_frames.append(block_df)

    if not new_frames:
        return display_df.copy()

    combined_new = pd.concat(new_frames, axis=1)

    if only_show_buckets:
        # Keep everything that ISN'T a raw slab / slab-subtotal column (i.e.
        # Grand Total, plus any other unrelated column), then splice the new
        # bucket block in right before Grand Total (if present).
        keep_tuples = [t for t in tuples if not _is_slab_related(t)]
        kept_df = display_df[keep_tuples] if keep_tuples else display_df.iloc[:, 0:0]
        grand_total_mask = [_is_grand_total(t) for t in keep_tuples]
        if any(grand_total_mask):
            gt_pos = grand_total_mask.index(True)
            before = kept_df.iloc[:, :gt_pos]
            after = kept_df.iloc[:, gt_pos:]
            result = pd.concat([before, combined_new, after], axis=1)
        else:
            result = pd.concat([kept_df, combined_new], axis=1)
        return result

    # Insert the new bucket blocks right before "Grand Total" (if present) so
    # Grand Total stays the visual right-most anchor; otherwise append at the end.
    grand_total_mask = [_is_grand_total(t) for t in tuples]
    if any(grand_total_mask):
        gt_pos = grand_total_mask.index(True)
        before = display_df.iloc[:, :gt_pos]
        after = display_df.iloc[:, gt_pos:]
        result = pd.concat([before, combined_new, after], axis=1)
    else:
        result = pd.concat([display_df, combined_new], axis=1)

    return result


# --------------------------------------------------------------------------- #
# General Field Efficiency Table: each cell as a % of the total across every
# OTHER value of a user-chosen field, holding every other selected Rows/
# Columns field fixed. Unlike a hardcoded "% of Stage" table, this works for
# *any* field the user points it at (a Rows field or a Columns field) — pick
# "STAGE_ECL" to get % of stage-group total per (Zone, Month); pick "zone"
# instead to get % of zone-group total per (Stage, Month); same formula.
# --------------------------------------------------------------------------- #
def _date_field_leaf(field: str, column_kind_map: dict[str, str]) -> str:
    """
    The Field Efficiency Table groups by whichever field the user picked at
    its full original granularity (matching its pre-optimization behaviour,
    which grouped directly on the raw date/timestamp column). Since raw date
    columns are no longer fetched at all -- only their SQL-bucketed Month/
    Date text columns are -- a date field maps here to its exact-Date leaf
    column ("__<field>__date"), which carries the same information (one
    entry per distinct date) the raw column did.
    """
    return f"__{field}__date" if column_kind_map.get(field) == "date" else field


@st.cache_data(ttl=300, show_spinner=False)
def build_field_efficiency_pivot(
    source_df: pd.DataFrame,
    row_cols: list[str],
    col_cols: list[str],
    normalize_field: str,
    value_col: str,
    agg_func: str,
    column_kind_map: dict[str, str],
) -> pd.DataFrame:
    """
    efficiency(cell) = value(cell) ÷ sum of value across every other value of
    `normalize_field`, holding every OTHER selected Rows/Columns field fixed.

    E.g. with Rows=[zone], Columns=[STAGE_ECL, value_date], normalize_field=
    "STAGE_ECL": Stage 1 / East / July = value(Stage1,East,July) ÷
    [value(Stage1,East,July) + value(Stage2,East,July) + value(Stage3,East,July)].
    Point it at "zone" instead and it normalizes across zones per (Stage, Month).

    Also appends a "Total" row, "Total" column, and "Total"/"Total" corner
    cell -- each computed the same correct way as every other cell: combine
    the underlying sum/count/min/max stats first at that (coarser) level,
    THEN derive the requested agg_func from them. Never averages the % cells
    that make it up, which would silently give the wrong number the moment
    group sizes differ (the classic "ratio of sums" vs "average of ratios"
    trap) -- and, for mean/min/max, never simply re-averages already
    per-cell-aggregated numbers either.

    `source_df` is the pre-aggregated stats table from `fetch_pivot_source_data`
    (one row per unique Rows x Columns combination, with STAT_SUM/STAT_COUNT/
    STAT_MIN/STAT_MAX columns), NOT raw per-record data -- this is what keeps
    the table fast to build even against a multi-million-row source table.

    Returns a DataFrame shaped like the raw pivot, plus that extra Total row
    and Total column, every cell already expressed as a % (×100).
    """
    field_row_cols = [_date_field_leaf(f, column_kind_map) for f in row_cols]
    field_col_cols = [_date_field_leaf(f, column_kind_map) for f in col_cols]
    field_normalize = _date_field_leaf(normalize_field, column_kind_map)

    all_dims = list(dict.fromkeys([*field_row_cols, *field_col_cols]))  # de-dup, preserve order
    if field_normalize not in all_dims:
        raise ValueError(f"'{normalize_field}' must be one of the selected Rows/Columns fields.")

    def _efficiency_series(dims: list[str]) -> pd.Series:
        """
        Combine the pre-aggregated stats over `dims` (a subset of all_dims,
        possibly empty) and express each entry as a % of the sum across
        every value of field_normalize, holding every OTHER dim in `dims`
        fixed. If field_normalize isn't part of `dims` at all -- i.e. it's
        already been collapsed away by moving to a coarser Total level --
        every entry is trivially its own 100% (0% if the whole slice is empty).
        """
        if not dims:
            total = _combine_group_stats(source_df, [], agg_func).iloc[0]
            return pd.Series([100.0 if total else 0.0], index=[0])

        leaf = _combine_group_stats(source_df, dims, agg_func)

        if field_normalize not in dims:
            return leaf.where(leaf == 0, 100.0).fillna(0.0)

        other = [d for d in dims if d != field_normalize]
        denom = leaf.groupby(level=other).transform("sum") if other else pd.Series(leaf.sum(), index=leaf.index)
        return (leaf.where(denom != 0) / denom.where(denom != 0) * 100.0).fillna(0.0)

    # --- Main cells (unchanged logic, now sourced from combined stats) ---
    cell_eff = _efficiency_series(all_dims)
    eff_flat = cell_eff.reset_index(name="efficiency_pct")
    pivot_eff = eff_flat.pivot_table(
        index=field_row_cols,
        columns=field_col_cols if field_col_cols else None,
        values="efficiency_pct",
        aggfunc="sum",
        fill_value=0.0,
    )
    original_columns = pivot_eff.columns  # capture before appending the Total column

    # --- Total column: collapse every Columns field, keep Rows as-is ---
    row_margin_eff = _efficiency_series(field_row_cols)
    total_col_key = ("Total",) + ("",) * (pivot_eff.columns.nlevels - 1) if isinstance(pivot_eff.columns, pd.MultiIndex) else "Total"
    pivot_eff[total_col_key] = row_margin_eff.reindex(pivot_eff.index).fillna(0.0).to_numpy()

    # --- Total row: collapse every Rows field, keep Columns as-is, plus the corner cell ---
    if field_col_cols:
        col_margin_eff = _efficiency_series(field_col_cols).reindex(original_columns).fillna(0.0)
    else:
        # No Columns field selected -> original_columns is a single synthetic
        # column ("efficiency_pct"), not something _efficiency_series([])
        # (indexed by [0]) can be reindexed against. Broadcast its one
        # trivial value across that single column instead.
        col_margin_eff = pd.Series([_efficiency_series([]).iloc[0]] * len(original_columns), index=original_columns)
    corner_value = _efficiency_series([]).iloc[0]

    total_row_index = pd.Index([("Total",) + ("",) * (pivot_eff.index.nlevels - 1)]) if isinstance(pivot_eff.index, pd.MultiIndex) else pd.Index(["Total"])
    total_row_values = list(col_margin_eff.to_numpy()) + [corner_value]
    total_row_df = pd.DataFrame([total_row_values], index=total_row_index, columns=pivot_eff.columns)

    pivot_eff = pd.concat([pivot_eff, total_row_df], axis=0)

    # Same chronological-month ordering as the Summary Matrix above, applied
    # here too since this table is built via groupby/pivot_table rather than
    # the row/column tree builder that already handles it.
    pivot_eff = pivot_eff.reindex(index=_reorder_axis_chronologically(pivot_eff.index))
    pivot_eff = pivot_eff.reindex(columns=_reorder_axis_chronologically(pivot_eff.columns))
    pivot_eff = _prettify_date_leaf_axes(pivot_eff, [*row_cols, *col_cols], column_kind_map)
    return pivot_eff


def _prettify_date_leaf_axes(
    df: pd.DataFrame, original_fields: list[str], column_kind_map: dict[str, str]
) -> pd.DataFrame:
    """
    Cosmetic pass for tables (like the Field Efficiency Table) built from the
    "__<field>__date" leaf column directly rather than through the row/column
    tree builder: renames that internal-looking axis name back to the
    original field name, and reformats its raw 'YYYY-MM-DD' values to
    'DD-Mon-YY' for display -- matching the date formatting used everywhere
    else in the app. Wrapped defensively: if anything about the axis shape
    is unexpected, the table is returned completely unchanged rather than
    risk breaking the view over a formatting nicety.
    """
    try:
        date_leaf_to_original = {
            f"__{f}__date": f for f in original_fields if column_kind_map.get(f) == "date"
        }
        if not date_leaf_to_original:
            return df

        def _relabel(idx: pd.Index) -> pd.Index:
            if isinstance(idx, pd.MultiIndex):
                new_names = [date_leaf_to_original.get(n, n) for n in idx.names]
                new_levels = []
                for level_i, name in enumerate(idx.names):
                    if name in date_leaf_to_original:
                        vals = idx.get_level_values(level_i)
                        parsed = pd.to_datetime(vals, format="%Y-%m-%d", errors="coerce")
                        pretty = parsed.strftime("%d-%b-%y")
                        new_levels.append([p if pd.notna(dt) else v for v, dt, p in zip(vals, parsed, pretty)])
                    else:
                        new_levels.append(list(idx.get_level_values(level_i)))
                return pd.MultiIndex.from_arrays(new_levels, names=new_names)
            if idx.name in date_leaf_to_original:
                parsed = pd.to_datetime(idx, format="%Y-%m-%d", errors="coerce")
                pretty = parsed.strftime("%d-%b-%y")
                new_vals = [p if pd.notna(dt) else v for v, dt, p in zip(idx, parsed, pretty)]
                return pd.Index(new_vals, name=date_leaf_to_original[idx.name])
            return idx

        df = df.copy()
        df.index = _relabel(df.index)
        df.columns = _relabel(df.columns)
        return df
    except Exception:
        return df


def _render_styled_efficiency_matrix(pivot_eff: pd.DataFrame, total_col_flags: list[bool]) -> None:
    """Shared styled st.dataframe rendering for the Field Efficiency Table
    (single- or multi-Values-column) -- bolds/shades the Total row and each
    block's own Total column, formats as %, falls back to a plain unstyled
    table if styling fails for any reason rather than crashing the page."""
    try:
        def _apply_total_styles(df: pd.DataFrame) -> pd.DataFrame:
            styles = pd.DataFrame("", index=df.index, columns=df.columns)
            last_row_label = df.index[-1]
            styles.loc[last_row_label, :] = "background-color:#dbe5f1; font-weight:bold"
            for j, is_total in enumerate(total_col_flags):
                if is_total:
                    for i in range(len(df)):
                        is_corner = df.index[i] == last_row_label
                        addition = "background-color:#c9d6ea; font-weight:bold" if is_corner else "background-color:#f2f2f2; font-weight:bold"
                        existing = styles.iloc[i, j]
                        styles.iloc[i, j] = f"{existing}; {addition}" if existing else addition
            return styles

        styler = pivot_eff.style.apply(_apply_total_styles, axis=None).format("{:,.2f}%")
        st.dataframe(styler, use_container_width=True)
    except Exception:
        st.dataframe(pivot_eff, use_container_width=True)


def render_field_efficiency_table(
    source_df: pd.DataFrame,
    row_cols: list[str],
    col_cols: list[str],
    normalize_field: str,
    value_col: str,
    agg_func: str,
    column_kind_map: dict[str, str],
) -> pd.DataFrame:
    """Build the general Field Efficiency table, render it, return it for CSV export."""
    pivot_eff = build_field_efficiency_pivot(
        source_df, row_cols, col_cols, normalize_field, value_col, agg_func, column_kind_map
    )

    st.markdown(f"**{agg_func.title()} of {value_col} — % of `{normalize_field}` group total**")
    st.caption(
        f"Each cell = its value ÷ the sum across every value of **`{normalize_field}`** for that same "
        "combination of the other selected Rows/Columns fields — computed automatically for every "
        "row and column in the table, however many there are. Pick a different field above to "
        "normalize a different way (e.g. % across zones instead of % across stages). The **Total** "
        "row/column/corner are computed the same correct way — the underlying sum/count/min/max are "
        "combined first, then the % is derived — never by averaging the % cells."
    )

    total_col_flags = [False] * (pivot_eff.shape[1] - 1) + [True]
    _render_styled_efficiency_matrix(pivot_eff, total_col_flags)
    return pivot_eff


def render_multi_value_field_efficiency_table(
    source_df_by_value_col: dict[str, pd.DataFrame],
    row_cols: list[str],
    col_cols: list[str],
    normalize_field: str,
    value_cols: list[str],
    agg_func: str,
    column_kind_map: dict[str, str],
) -> pd.DataFrame:
    """
    Same idea as `render_multi_value_pivot_table`, applied to the Field
    Efficiency Table: build the IDENTICAL % table (same Rows/Columns/
    normalize_field -- only the underlying numbers differ) once per Values
    column, using the exact same tested single-value engine, then place the
    results side by side with an extra outer header level naming each one.
    Every block is reindexed to the FIRST block's row order before
    concatenating, same safeguard as the Summary Matrix version.
    """
    blocks: dict[str, pd.DataFrame] = {}
    total_col_flags: list[bool] = []
    base_index: pd.Index | None = None

    for vc in value_cols:
        block_df = build_field_efficiency_pivot(
            source_df_by_value_col[vc], row_cols, col_cols, normalize_field, vc, agg_func, column_kind_map
        )
        if base_index is None:
            base_index = block_df.index
        else:
            block_df = block_df.reindex(base_index).fillna(0.0)
        blocks[f"{agg_func.title()} of {vc}"] = block_df
        total_col_flags.extend([False] * (block_df.shape[1] - 1) + [True])

    pivot_eff = pd.concat(blocks, axis=1)

    st.markdown(
        f"**{agg_func.title()} of {len(value_cols)} Values columns — % of `{normalize_field}` group total** "
        "— one column block per field, side by side"
    )
    st.caption(
        f"Each cell = its value ÷ the sum across every value of **`{normalize_field}`** for that same "
        "combination of the other selected Rows/Columns fields — computed separately for each Values "
        "column, then placed side by side. Never averaged or re-derived across columns."
    )
    _render_styled_efficiency_matrix(pivot_eff, total_col_flags)
    return pivot_eff



def _clean_efficiency_column_label(label: str) -> str:
    """'Sum of march_23' -> 'march_23', 'Mean of april_23' -> 'april_23', etc. Leaves anything else untouched."""
    for agg_word in ("Sum", "Mean", "Count", "Min", "Max"):
        prefix = f"{agg_word} of "
        if label.startswith(prefix):
            return label[len(prefix):]
    return label


def slice_field_efficiency_to_status(
    pivot_eff: pd.DataFrame,
    normalize_field: str,
    status_value: str,
) -> pd.DataFrame:
    """
    Slice an ALREADY-COMPUTED Field Efficiency Table down to just one value
    of `normalize_field` for display (e.g. only 'BOUNCED'), dropping every
    other status value, the 'Total' column, and the now-redundant
    normalize_field header level. Also strips any leading 'Sum of ' /
    'Mean of ' etc. from the remaining Values-column labels.

    CRITICAL: `pivot_eff` must have been built by build_field_efficiency_pivot
    over the FULL dataset (every status value, e.g. BOUNCED *and* CLEARED)
    BEFORE calling this. This function only slices the finished % table for
    display -- it never touches the computation. Filtering the source data
    to just BOUNCED before computing would collapse the denominator to
    BOUNCED/BOUNCED = 100%, which is exactly what this two-step order avoids.
    """
    cols = pivot_eff.columns

    if isinstance(cols, pd.MultiIndex):
        level_names = list(cols.names)
        level = (
            level_names.index(normalize_field)
            if normalize_field in level_names
            else next((i for i in range(cols.nlevels) if status_value in cols.get_level_values(i)), None)
        )
        if level is None:
            raise ValueError(f"Could not find '{status_value}' in any column level of {level_names}.")
        sliced = pivot_eff.xs(status_value, axis=1, level=level, drop_level=True)
    else:
        if status_value not in cols:
            raise ValueError(f"'{status_value}' not found in columns: {list(cols)}")
        sliced = pivot_eff[[status_value]]

    sliced = sliced.copy()
    if isinstance(sliced.columns, pd.MultiIndex):
        sliced.columns = pd.MultiIndex.from_tuples(
            tuple(_clean_efficiency_column_label(v) if isinstance(v, str) else v for v in tup)
            for tup in sliced.columns
        )
    else:
        sliced.columns = [_clean_efficiency_column_label(c) if isinstance(c, str) else c for c in sliced.columns]

    return sliced


def build_pivot_table(
    agg_df: pd.DataFrame,
    row_cols: list[str],
    col_cols: list[str],
    agg_func: str,
) -> pd.DataFrame:
    """
    Build a flat pivot matrix (single overall Grand Total row/column, no
    per-level subtotals) directly from the pre-aggregated stats table --
    used for the Chart and the flat CSV download. `row_cols`/`col_cols` here
    are the *effective* (already date-expanded) field names, i.e. exactly
    the columns present in `agg_df`.

    Grand-total margins are reconstructed CORRECTLY for every agg_func
    (sum/count/mean/min/max) by re-combining the underlying sum/count/min/
    max stats via `_combine_group_stats` -- never by summing already
    per-cell-aggregated values, which is exactly right for sum/count but
    silently wrong for mean/min/max (pandas' own `pivot_table(margins=True)`
    makes this mistake for anything other than sum/count).
    """
    leaf = _combine_group_stats(agg_df, [*row_cols, *col_cols], agg_func)
    leaf_flat = leaf.reset_index(name="value")
    pivot = leaf_flat.pivot_table(
        index=row_cols, columns=col_cols if col_cols else None, values="value", aggfunc="sum", fill_value=0,
    )
    # aggfunc="sum" above is inert/safe: `leaf` already has exactly one row
    # per unique Rows x Columns combination (the full grouping granularity),
    # so this pivot_table call only *reshapes* -- it never combines more
    # than one raw value into a cell.
    original_columns = pivot.columns

    row_total = _combine_group_stats(agg_df, row_cols, agg_func)
    total_col_key = ("Total",) + ("",) * (pivot.columns.nlevels - 1) if isinstance(pivot.columns, pd.MultiIndex) else "Total"
    pivot[total_col_key] = row_total.reindex(pivot.index).fillna(0.0).to_numpy()

    if col_cols:
        col_total = _combine_group_stats(agg_df, col_cols, agg_func).reindex(original_columns).fillna(0.0)
    else:
        col_total = pd.Series([_combine_group_stats(agg_df, [], agg_func).iloc[0]] * len(original_columns), index=original_columns)
    corner = _combine_group_stats(agg_df, [], agg_func).iloc[0]

    if isinstance(pivot.index, pd.MultiIndex):
        total_row_index = pd.MultiIndex.from_tuples(
            [("Total",) + ("",) * (pivot.index.nlevels - 1)], names=pivot.index.names
        )
    else:
        # Preserve the original index name (e.g. "del_month_name") explicitly --
        # pd.concat sets the resulting index name to None whenever the pieces
        # being concatenated disagree on it, and an unnamed pd.Index(["Total"])
        # always disagrees with the real (named) pivot index. Losing the name
        # here breaks every downstream reset_index()/melt() call that expects
        # a column literally called by that field's name (e.g. the Chart).
        total_row_index = pd.Index(["Total"], name=pivot.index.name)
    total_row_values = list(col_total.to_numpy()) + [corner]
    total_row_df = pd.DataFrame([total_row_values], index=total_row_index, columns=pivot.columns)

    pivot = pd.concat([pivot, total_row_df], axis=0)
    pivot = pivot.reindex(index=_reorder_axis_chronologically(pivot.index))
    pivot = pivot.reindex(columns=_reorder_axis_chronologically(pivot.columns))
    return pivot


def _drop_pivot_totals(pivot_df: pd.DataFrame) -> pd.DataFrame:
    """Strip the 'Total' margin row/column so charts show only real categories."""
    trimmed = pivot_df.copy()
    if "Total" in trimmed.index:
        trimmed = trimmed.drop(index="Total")
    if isinstance(trimmed.columns, pd.MultiIndex):
        mask = trimmed.columns.get_level_values(0) != "Total"
        trimmed = trimmed.loc[:, mask]
    elif "Total" in trimmed.columns:
        trimmed = trimmed.drop(columns="Total")
    return trimmed


def render_pivot_chart(
    pivot_df: pd.DataFrame,
    row_cols: list[str],
    col_cols: list[str],
    value_col: str,
    agg_func: str,
    key_suffix: str,
    display_row_cols: list[str] | None = None,
    display_col_cols: list[str] | None = None,
) -> None:
    """
    Render an interactive Plotly grouped bar chart matching the pivoted
    dimensions. `row_cols`/`col_cols` must be the *effective* field names
    (matching `pivot_df`'s actual index/column names -- a date field
    expands to "__field__month"/"__field__date"). `display_row_cols`/
    `display_col_cols` are the original, user-facing field names shown in
    the title/legend; they default to `row_cols`/`col_cols` when omitted.
    """
    display_row_cols = display_row_cols if display_row_cols is not None else row_cols
    display_col_cols = display_col_cols if display_col_cols is not None else col_cols

    chart_df = _drop_pivot_totals(pivot_df)
    if chart_df.empty:
        st.info("Nothing to chart once the grand-total row/column is excluded.")
        return

    # Flatten a MultiIndex column axis (happens when 2+ "Columns" fields are chosen)
    # into single readable string labels before melting for Plotly.
    if isinstance(chart_df.columns, pd.MultiIndex):
        chart_df.columns = [
            " | ".join(str(part) for part in col if str(part) != "") if isinstance(col, tuple) else str(col)
            for col in chart_df.columns
        ]

    chart_df = chart_df.reset_index()
    value_vars = [c for c in chart_df.columns if c not in row_cols]
    if not value_vars:
        st.info("No values available to chart.")
        return

    long_df = chart_df.melt(id_vars=row_cols, value_vars=value_vars, var_name="__pivot_col__", value_name="__value__")

    if len(row_cols) == 1:
        x_col = row_cols[0]
    else:
        x_col = "__row_label__"
        long_df[x_col] = long_df[row_cols].astype(str).agg(" | ".join, axis=1)

    color_col = "__pivot_col__" if col_cols else None
    title_bits = f"{agg_func.title()} of {value_col} by {', '.join(display_row_cols)}"
    if display_col_cols:
        title_bits += f", split by {', '.join(display_col_cols)}"

    fig = px.bar(
        long_df,
        x=x_col,
        y="__value__",
        color=color_col,
        barmode="group",
        template="plotly_white",
        title=title_bits,
        labels={
            "__value__": f"{agg_func}({value_col})",
            x_col: " | ".join(display_row_cols),
            "__pivot_col__": " | ".join(display_col_cols) if display_col_cols else "",
        },
    )
    fig.update_layout(
        height=550,
        font=dict(size=13),
        title_font_size=16,
        margin=dict(t=70, l=50, r=30, b=60),
        legend_title_text=" | ".join(display_col_cols) if display_col_cols else "",
    )
    config = {"displaylogo": False, "toImageButtonOptions": {"format": "png", "filename": f"pivot_chart_{key_suffix}", "scale": 3}}
    st.plotly_chart(fig, use_container_width=True, config=config, key=f"pivot_chart_{key_suffix}")


# --------------------------------------------------------------------------- #
# Authentication & Authorization
# --------------------------------------------------------------------------- #
# Two tables, on purpose:
#   authorized_employees -- WHO is allowed to have an account at all (admin-managed)
#   app_users            -- the password hash for employees who created one
# Kept separate so access can be revoked (is_authorized = FALSE) without
# touching the password, and so "authorized" and "has an account" are two
# independently-true-or-false things.
_AUTH_SESSION_KEYS = ("auth_employee_code", "auth_employee_name", "auth_access_level")

_CREATE_EMPLOYEES_SQL = text(
    """
    CREATE TABLE IF NOT EXISTS authorized_employees (
        employee_code   VARCHAR(20) PRIMARY KEY,
        employee_name   VARCHAR(100) NOT NULL,
        department      VARCHAR(100),
        access_level    VARCHAR(20) NOT NULL DEFAULT 'user'
                        CHECK (access_level IN ('admin', 'user', 'analyst')),
        is_authorized   BOOLEAN NOT NULL DEFAULT TRUE,
        created_at      TIMESTAMPTZ NOT NULL DEFAULT now()
    )
    """
)
_CREATE_USERS_SQL = text(
    """
    CREATE TABLE IF NOT EXISTS app_users (
        employee_code   VARCHAR(20) PRIMARY KEY
                        REFERENCES authorized_employees(employee_code) ON DELETE CASCADE,
        password_hash   TEXT NOT NULL,
        created_at      TIMESTAMPTZ NOT NULL DEFAULT now(),
        last_login      TIMESTAMPTZ
    )
    """
)
_CREATE_AUDIT_SQL = text(
    """
    CREATE TABLE IF NOT EXISTS login_audit (
        id              BIGSERIAL PRIMARY KEY,
        employee_code   VARCHAR(20) NOT NULL,
        event           VARCHAR(20) NOT NULL
                        CHECK (event IN ('login_success', 'login_failure', 'logout', 'account_created')),
        event_time      TIMESTAMPTZ NOT NULL DEFAULT now()
    )
    """
)


@st.cache_resource(show_spinner=False)
def ensure_auth_tables(_engine: Engine) -> None:
    """Create the auth tables if they don't exist yet. Runs once per app process."""
    with _engine.begin() as conn:
        conn.execute(_CREATE_EMPLOYEES_SQL)
        conn.execute(_CREATE_USERS_SQL)
        conn.execute(_CREATE_AUDIT_SQL)


def _is_employee_authorized(engine: Engine, employee_code: str) -> tuple[bool, str | None, str | None]:
    """Return (is_authorized, employee_name, access_level); (False, None, None) if unknown/disabled."""
    query = text(
        "SELECT employee_name, access_level, is_authorized "
        "FROM authorized_employees WHERE employee_code = :code"
    )
    with engine.connect() as conn:
        row = conn.execute(query, {"code": employee_code}).mappings().fetchone()
    if row is None or not row["is_authorized"]:
        return False, None, None
    return True, row["employee_name"], row["access_level"]


def _get_password_hash(engine: Engine, employee_code: str) -> str | None:
    query = text("SELECT password_hash FROM app_users WHERE employee_code = :code")
    with engine.connect() as conn:
        row = conn.execute(query, {"code": employee_code}).fetchone()
    return row[0] if row else None


def _account_exists(engine: Engine, employee_code: str) -> bool:
    return _get_password_hash(engine, employee_code) is not None


def _create_account(engine: Engine, employee_code: str, password: str) -> None:
    password_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    query = text("INSERT INTO app_users (employee_code, password_hash) VALUES (:code, :hash)")
    with engine.begin() as conn:
        conn.execute(query, {"code": employee_code, "hash": password_hash})
    _log_audit_event(engine, employee_code, "account_created")


def _verify_password(password: str, password_hash: str) -> bool:
    try:
        return bcrypt.checkpw(password.encode("utf-8"), password_hash.encode("utf-8"))
    except (ValueError, TypeError):
        return False


def _touch_last_login(engine: Engine, employee_code: str) -> None:
    query = text("UPDATE app_users SET last_login = now() WHERE employee_code = :code")
    with engine.begin() as conn:
        conn.execute(query, {"code": employee_code})


def _log_audit_event(engine: Engine, employee_code: str, event: str) -> None:
    query = text("INSERT INTO login_audit (employee_code, event) VALUES (:code, :event)")
    try:
        with engine.begin() as conn:
            conn.execute(query, {"code": employee_code, "event": event})
    except SQLAlchemyError:
        pass  # audit logging must never be the reason a login fails


def _too_many_recent_failures(
    engine: Engine, employee_code: str, max_attempts: int = 5, window_minutes: int = 15
) -> bool:
    """Basic brute-force throttle: block further attempts if too many recent failures.

    Checked against the database (not st.session_state) so it can't be bypassed
    by simply reloading the page or opening a new browser tab.
    """
    query = text(
        "SELECT COUNT(*) FROM login_audit "
        "WHERE employee_code = :code AND event = 'login_failure' "
        "AND event_time > now() - (:window || ' minutes')::interval"
    )
    with engine.connect() as conn:
        count = conn.execute(query, {"code": employee_code, "window": window_minutes}).scalar()
    return (count or 0) >= max_attempts


def _render_login_form(engine: Engine) -> None:
    with st.form("login_form"):
        employee_code = st.text_input("Employee Code").strip().upper()
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login", use_container_width=True)

    if not submitted:
        return
    if not employee_code or not password:
        st.error("Please enter both your Employee Code and Password.")
        return
    if _too_many_recent_failures(engine, employee_code):
        st.error("Too many failed attempts for this employee code. Please try again in a few minutes.")
        return

    authorized, employee_name, access_level = _is_employee_authorized(engine, employee_code)
    password_hash = _get_password_hash(engine, employee_code)

    # Deliberately one generic error for every failure reason (not authorized /
    # no account yet / wrong password) -- so a login attempt never reveals
    # which employee codes exist, are authorized, or have registered.
    generic_error = "Invalid employee code or password."
    if not authorized or password_hash is None or not _verify_password(password, password_hash):
        _log_audit_event(engine, employee_code, "login_failure")
        st.error(generic_error)
        return

    _touch_last_login(engine, employee_code)
    _log_audit_event(engine, employee_code, "login_success")
    st.session_state["auth_employee_code"] = employee_code
    st.session_state["auth_employee_name"] = employee_name
    st.session_state["auth_access_level"] = access_level
    st.rerun()


def _render_create_account_form(engine: Engine) -> None:
    st.caption("Only employee codes explicitly authorized by your admin can create an account.")
    with st.form("create_account_form"):
        employee_code = st.text_input("Employee Code", key="create_code").strip().upper()
        password = st.text_input("Password", type="password", key="create_pw")
        confirm_password = st.text_input("Confirm Password", type="password", key="create_pw2")
        submitted = st.form_submit_button("Create Account", use_container_width=True)

    if not submitted:
        return
    if not employee_code or not password:
        st.error("Please fill in all fields.")
        return
    if password != confirm_password:
        st.error("Passwords do not match.")
        return
    if len(password) < 8:
        st.error("Please choose a password with at least 8 characters.")
        return

    authorized, employee_name, _access_level = _is_employee_authorized(engine, employee_code)
    if not authorized:
        st.error("Your employee code is not authorized to access this application. Please contact your administrator.")
        return
    if _account_exists(engine, employee_code):
        st.error("An account already exists for this employee code. Please log in instead.")
        return

    _create_account(engine, employee_code, password)
    st.success(f"Account created for {employee_name or employee_code}. You can now log in from the Login tab.")


def require_authentication(engine: Engine) -> None:
    """Gate the rest of the app behind login. Stops execution until authenticated."""
    if st.session_state.get("auth_employee_code"):
        return  # already logged in this session

    st.title("🔒 Employee Login")
    st.caption("This application is restricted to authorized department employees.")

    login_tab, create_tab = st.tabs(["Login", "Create Account"])
    with login_tab:
        _render_login_form(engine)
    with create_tab:
        _render_create_account_form(engine)

    st.stop()


def render_sidebar_identity(engine: Engine) -> None:
    """Small identity block + logout button, shown at the top of the sidebar."""
    name = st.session_state.get("auth_employee_name") or st.session_state["auth_employee_code"]
    level = st.session_state.get("auth_access_level", "user")
    st.markdown(f"👤 **{name}**  \n`{st.session_state['auth_employee_code']}` · {level}")
    if st.button("Logout", use_container_width=True):
        _log_audit_event(engine, st.session_state["auth_employee_code"], "logout")
        for key in _AUTH_SESSION_KEYS:
            st.session_state.pop(key, None)
        st.rerun()
    st.divider()


def render_admin_panel(engine: Engine) -> None:
    """Admin-only: manage the authorized-employee list. Caller must check role first."""
    st.subheader("🛡️ Authorized Employees")
    st.caption("Only employees listed here (and authorized) can create an account or log in.")

    query = text(
        "SELECT employee_code, employee_name, department, access_level, is_authorized "
        "FROM authorized_employees ORDER BY employee_code"
    )
    with engine.connect() as conn:
        employees_df = pd.read_sql(query, conn)
    st.dataframe(employees_df, use_container_width=True, height=300)

    st.markdown("#### ➕ Add an authorized employee")
    with st.form("admin_add_employee"):
        col1, col2 = st.columns(2)
        with col1:
            new_code = st.text_input("Employee Code").strip().upper()
            new_name = st.text_input("Employee Name").strip()
        with col2:
            new_dept = st.text_input("Department").strip()
            new_level = st.selectbox("Access Level", ["user", "analyst", "admin"])
        add_submitted = st.form_submit_button("Add Employee", use_container_width=True)

    if add_submitted:
        if not new_code or not new_name:
            st.error("Employee Code and Employee Name are required.")
        else:
            upsert_query = text(
                """
                INSERT INTO authorized_employees (employee_code, employee_name, department, access_level, is_authorized)
                VALUES (:code, :name, :dept, :level, TRUE)
                ON CONFLICT (employee_code) DO UPDATE
                SET employee_name = EXCLUDED.employee_name,
                    department = EXCLUDED.department,
                    access_level = EXCLUDED.access_level,
                    is_authorized = TRUE
                """
            )
            with engine.begin() as conn:
                conn.execute(upsert_query, {"code": new_code, "name": new_name, "dept": new_dept, "level": new_level})
            st.success(f"{new_code} added / re-authorized.")
            st.rerun()

    st.markdown("#### 🔁 Enable / disable an employee")
    if not employees_df.empty:
        toggle_code = st.selectbox("Employee Code", employees_df["employee_code"].tolist(), key="admin_toggle_code")
        current_row = employees_df.loc[employees_df["employee_code"] == toggle_code].iloc[0]
        col_a, col_b = st.columns(2)
        with col_a:
            if current_row["is_authorized"] and st.button("🚫 Disable", use_container_width=True):
                with engine.begin() as conn:
                    conn.execute(
                        text("UPDATE authorized_employees SET is_authorized = FALSE WHERE employee_code = :code"),
                        {"code": toggle_code},
                    )
                st.success(f"{toggle_code} disabled.")
                st.rerun()
        with col_b:
            if not current_row["is_authorized"] and st.button("✅ Re-enable", use_container_width=True):
                with engine.begin() as conn:
                    conn.execute(
                        text("UPDATE authorized_employees SET is_authorized = TRUE WHERE employee_code = :code"),
                        {"code": toggle_code},
                    )
                st.success(f"{toggle_code} re-enabled.")
                st.rerun()

    st.markdown("#### 📜 Recent login activity")
    audit_query = text(
        "SELECT employee_code, event, event_time FROM login_audit ORDER BY event_time DESC LIMIT 50"
    )
    with engine.connect() as conn:
        audit_df = pd.read_sql(audit_query, conn)
    st.dataframe(audit_df, use_container_width=True, height=250)


# --------------------------------------------------------------------------- #
# File Upload -> Database (admin / analyst only)
# --------------------------------------------------------------------------- #
# Ported from the standalone upload.py script, with one important fix: date
# columns are converted to real `date` objects (via .dt.date) and written
# with an explicit SQL DATE dtype, so Postgres stores them with no 00:00:00
# time component -- instead of drifting into a TIMESTAMP with a stray time.
_DATE_KEYWORDS = ("date", "month", "year", "time", "dt", "period")


def _format_date_like_header(col: Any) -> str:
    """Catch a date-like column header (real datetime, or date-looking string)
    and format it as e.g. 'March 23' before it gets snake_cased."""
    if hasattr(col, "strftime"):
        return col.strftime("%B %y")
    if isinstance(col, str) and ("/" in col or "-" in col or "00:00" in col):
        try:
            dt = pd.to_datetime(col, dayfirst=True)
            return dt.strftime("%B %y")
        except Exception:
            return col
    return str(col)


def _clean_column_name(name: str) -> str:
    name = str(name).strip().lower().replace(" ", "_")
    return re.sub(r"[^a-zA-Z0-9_]", "", name)


def _dedupe_columns(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for col in columns:
        if col in seen:
            seen[col] += 1
            result.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            result.append(col)
    return result


def _excel_engine_for(filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext == "xlsb":
        return "pyxlsb"
    if ext == "xls":
        return "xlrd"
    return "openpyxl"  # .xlsx / .xlsm


def process_uploaded_workbook(uploaded_file) -> tuple[pd.DataFrame, list[str]]:
    """
    Read every sheet, clean + snake_case headers (catching date-like headers
    first, e.g. a real Timestamp column header becomes 'march_23'), combine
    all sheets, dedupe column names, and convert any date/month/year/time/dt
    /period-named column into a true `date` (no time component).

    Returns (combined_dataframe, list_of_columns_converted_to_date).
    """
    engine_name = _excel_engine_for(uploaded_file.name)
    all_sheets = pd.read_excel(uploaded_file, sheet_name=None, engine=engine_name)

    cleaned_sheets = []
    for _sheet_name, sheet_df in all_sheets.items():
        sheet_df = sheet_df.copy()
        sheet_df.columns = [_format_date_like_header(col) for col in sheet_df.columns]
        sheet_df.columns = [_clean_column_name(col) for col in sheet_df.columns]
        cleaned_sheets.append(sheet_df)

    df = pd.concat(cleaned_sheets, ignore_index=True)
    df.columns = _dedupe_columns(list(df.columns))

    date_cols_converted: list[str] = []
    for col in df.columns:
        if any(kw in col.lower() for kw in _DATE_KEYWORDS):
            if pd.api.types.is_numeric_dtype(df[col]):
                # Excel's serial date epoch (day 0 = 30-Dec-1899)
                parsed = pd.to_datetime(df[col], unit="D", origin="1899-12-30", errors="coerce")
            else:
                # format="mixed" parses each value independently -- without it,
                # pandas guesses ONE format from the first value and silently
                # turns every differently-formatted date in the same column
                # (e.g. some rows with a trailing 00:00:00, some without) into
                # a blank instead of the real date.
                parsed = pd.to_datetime(df[col], dayfirst=True, format="mixed", errors="coerce")
            df[col] = parsed.dt.date  # strip the time component entirely -- no more 00:00:00
            date_cols_converted.append(col)

    return df, date_cols_converted


def render_upload_tab(engine: Engine) -> None:
    st.subheader("📤 Upload Data to Database")
    st.caption("Upload an Excel file — every sheet is cleaned, combined, and written as one table.")

    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xlsm", "xlsb", "xls"])
    if uploaded_file is not None:
        try:
            with st.spinner("Reading and cleaning the workbook..."):
                df, date_cols = process_uploaded_workbook(uploaded_file)
        except ImportError as exc:
            st.error(
                f"❌ Missing the Excel engine needed to read this file type ({exc}). "
                "Ask your admin to run `pip install pyxlsb xlrd` on the server for "
                ".xlsb / .xls support — .xlsx/.xlsm work out of the box."
            )
            return
        except Exception as exc:  # noqa: BLE001
            st.error(f"❌ Failed to read the uploaded file: {exc}")
            return

        st.success(f"Loaded **{len(df):,}** rows and **{len(df.columns)}** columns from **{uploaded_file.name}**.")
        if date_cols:
            st.caption(f"🗓️ Cleaned as pure dates (no time component): {', '.join(date_cols)}")
        st.dataframe(df.head(20), use_container_width=True)

        default_table_name = _clean_column_name(uploaded_file.name.rsplit(".", 1)[0])
        col1, col2 = st.columns([2, 1])
        with col1:
            table_name = st.text_input("Destination table name", value=default_table_name)
        with col2:
            if_exists = st.selectbox("If table already exists", ["replace", "append", "fail"])

        if st.button("⬆️ Upload to Database", type="primary", use_container_width=True):
            table_name_clean = _clean_column_name(table_name)
            if not table_name_clean:
                st.error("Please enter a valid table name.")
            else:
                dtype_map = {col: sa_types.Date() for col in date_cols}
                try:
                    with st.spinner(f"Writing to table `{table_name_clean}`..."):
                        df.to_sql(
                            table_name_clean, engine, if_exists=if_exists, index=False,
                            chunksize=1000, dtype=dtype_map,
                        )
                    st.success(f"✅ Uploaded to table `{table_name_clean}` ({len(df):,} rows).")
                    list_tables.clear()
                except SQLAlchemyError as exc:
                    st.error(f"❌ Database error while uploading: {exc}")
                except Exception as exc:  # noqa: BLE001
                    st.error(f"❌ Unexpected error while uploading: {exc}")

    st.divider()
    st.markdown("#### 🗓️ Auto-generate `del_month_name` / `del_year`")
    st.caption(
        "Replaces the manual UPDATE query: pick a table and its date column, and this "
        "adds two columns — the month name only, and the year only — computed in PostgreSQL."
    )
    try:
        available_tables = list_tables(engine)
    except Exception as exc:  # noqa: BLE001
        st.error(f"❌ Could not list tables: {exc}")
        return
    if not available_tables:
        st.info("No tables found yet — upload a file above first.")
        return

    target_table = st.selectbox("Table", available_tables, key="del_cols_table")
    try:
        col_meta = get_columns(engine, target_table)
    except Exception as exc:  # noqa: BLE001
        st.error(f"❌ Could not list columns for `{target_table}`: {exc}")
        return

    date_like_cols = [c["name"] for c in col_meta if classify_column(c["type"]) == "date"]
    ref_col_options = date_like_cols if date_like_cols else [c["name"] for c in col_meta]
    if not date_like_cols:
        st.warning(
            "No columns in this table are stored as a real DATE/TIMESTAMP type yet — "
            "showing all columns, but this only works correctly on a genuine date column "
            "(exactly what a fresh upload above produces)."
        )
    ref_col = st.selectbox("Date column to derive month/year from", ref_col_options, key="del_cols_ref")

    if st.button("🧮 Generate del_month_name & del_year", use_container_width=True):
        try:
            with engine.begin() as conn:
                conn.execute(text(f'ALTER TABLE "{target_table}" ADD COLUMN IF NOT EXISTS del_month_name VARCHAR(20)'))
                conn.execute(text(f'ALTER TABLE "{target_table}" ADD COLUMN IF NOT EXISTS del_year INTEGER'))
                conn.execute(text(
                    f'UPDATE "{target_table}" SET '
                    f'del_month_name = TO_CHAR("{ref_col}"::date, \'FMMonth\'), '
                    f'del_year = EXTRACT(YEAR FROM "{ref_col}"::date)::INT'
                ))
            st.success(f"✅ `del_month_name` and `del_year` added to `{target_table}`, derived from `{ref_col}`.")
            get_columns.clear()
        except SQLAlchemyError as exc:
            st.error(f"❌ Database error: {exc}")
        except Exception as exc:  # noqa: BLE001
            st.error(f"❌ Unexpected error: {exc}")


# --------------------------------------------------------------------------- #
# UI - Sidebar: connection status
# --------------------------------------------------------------------------- #
st.title("📊 Dynamic Database Explorer & Report Generator")
st.caption("Select a table, filter by any column, and export a custom report.")

engine = get_engine_safe()
if engine is None:
    st.info(
        "Set your database credentials as environment variables (or in "
        "`.streamlit/secrets.toml`) and reload the app. See the README / "
        "setup instructions for details."
    )
    st.stop()

ensure_auth_tables(engine)
require_authentication(engine)  # stops execution here until logged in

with st.sidebar:
    render_sidebar_identity(engine)
    st.header("⚙️ Configuration")
    st.success(f"Connected to **{DB_NAME}**@`{DB_HOST}`", icon="✅")

    try:
        tables = list_tables(engine)
    except SQLAlchemyError as exc:
        st.error(f"Failed to list tables: {exc}")
        st.stop()

    if not tables:
        st.warning("No tables found in the `public` schema.")
        st.stop()

    st.divider()
    selected_table = st.selectbox("📁 Select a table", options=tables, index=0)

# --------------------------------------------------------------------------- #
# Load column metadata for the selected table
# --------------------------------------------------------------------------- #
try:
    columns_meta = get_columns(engine, selected_table)
except SQLAlchemyError as exc:
    st.error(f"Failed to fetch columns for `{selected_table}`: {exc}")
    st.stop()

column_names = [c["name"] for c in columns_meta]
column_kind_map = {c["name"]: classify_column(c["type"]) for c in columns_meta}

with st.sidebar:
    st.divider()
    st.subheader("1️⃣ Reference (Filter) Column")
    ref_column = st.selectbox("Reference column", options=column_names, key="ref_col")

    detected_kind = column_kind_map.get(ref_column, "text")
    kind_override = st.radio(
        "Treat this column as",
        options=["Date/Timestamp", "Categorical/Text"],
        index=0 if detected_kind == "date" else 1,
        help="Auto-detected from the database column type; override if needed.",
    )
    ref_kind = "date" if kind_override == "Date/Timestamp" else "text"

    filter_payload: dict[str, Any] = {}

    if ref_kind == "date":
        granularity = st.radio(
            "Granularity",
            options=["Specific Date(s)", "Month/Year", "Year(s)"],
            key="granularity",
        )
        filter_payload["granularity"] = granularity

        if granularity == "Specific Date(s)":
            picked_dates = st.date_input(
                "Pick one or more dates",
                value=[],
                help="Click multiple dates; use the calendar's range/multi picker.",
            )
            # st.date_input returns a single date, a tuple, or a list depending
            # on interaction state -- normalize to a list of dates.
            if isinstance(picked_dates, (list, tuple)):
                filter_payload["dates"] = list(picked_dates)
            elif picked_dates:
                filter_payload["dates"] = [picked_dates]
            else:
                filter_payload["dates"] = []

        elif granularity == "Month/Year":
            try:
                available_ym = get_available_year_months(engine, selected_table, ref_column)
            except SQLAlchemyError as exc:
                st.error(f"Failed to fetch available months: {exc}")
                available_ym = []
            ym_labels = {f"{MONTH_NAMES[mo - 1]} {yr}": (yr, mo) for yr, mo in available_ym}
            selected_labels = st.multiselect("Select Month/Year", options=list(ym_labels.keys()))
            filter_payload["year_months"] = [ym_labels[lbl] for lbl in selected_labels]

        elif granularity == "Year(s)":
            try:
                available_years = get_available_years(engine, selected_table, ref_column)
            except SQLAlchemyError as exc:
                st.error(f"Failed to fetch available years: {exc}")
                available_years = []
            filter_payload["years"] = st.multiselect("Select Year(s)", options=available_years)

    else:
        try:
            distinct_values = get_distinct_values(engine, selected_table, ref_column)
        except SQLAlchemyError as exc:
            st.error(f"Failed to fetch distinct values: {exc}")
            distinct_values = []
        filter_payload["values"] = st.multiselect(
            f"Select value(s) of `{ref_column}`", options=distinct_values
        )

    st.divider()
    st.subheader("2️⃣ Target / Output Columns")
    output_candidates = [c for c in column_names if c != ref_column]
    output_columns = st.multiselect(
        "Select one or more output columns",
        options=output_candidates,
        default=output_candidates[:1] if output_candidates else [],
    )

    st.divider()
    st.subheader("3️⃣ Monthly Cumulative Summary (optional)")
    show_monthly_summary = st.checkbox(
        "Show monthly cumulative summary",
        value=False,
        help=(
            "Counts records per calendar month — a record dated 01-03-2024 is "
            "counted as a March record. Optionally break it down by another "
            "column, e.g. payment_mode (Cash / Cheque / DD)."
        ),
    )

    summary_date_col: str | None = None
    breakdown_col: str | None = None
    focus_year_months: list[tuple[int, int]] = []
    focus_years_summary: list[int] = []
    focus_categories: list[Any] = []
    chart_res_scale: int = 3
    sum_columns: list[str] = []
    show_in_crores: bool = False
    num_chart_splits: int = 1

    if show_monthly_summary:
        date_columns = [c for c in column_names if column_kind_map.get(c) == "date"]
        if not date_columns:
            st.warning("No date/timestamp columns found in this table — can't build a monthly summary.")
            show_monthly_summary = False
        else:
            default_date_col = ref_column if ref_kind == "date" and ref_column in date_columns else date_columns[0]
            summary_date_col = st.selectbox(
                "Date column to summarize by month",
                options=date_columns,
                index=date_columns.index(default_date_col),
                key=f"summary_date_col_{selected_table}",
            )

            breakdown_candidates = ["(None)"] + [c for c in column_names if c != summary_date_col]
            breakdown_choice = st.selectbox(
                "Breakdown by (optional) — e.g. payment_mode, or a numeric column like emi to sum it",
                options=breakdown_candidates,
                key=f"breakdown_choice_{selected_table}",
                help=(
                    "Pick a text/category column (e.g. payment_mode) to split the chart "
                    "into groups. Pick a numeric column (e.g. emi) to total it instead — "
                    "numeric columns are never used to split/color the chart, since they "
                    "usually have far too many distinct values to group by."
                ),
            )

            if breakdown_choice == "(None)":
                breakdown_col = None
            elif column_kind_map.get(breakdown_choice) == "numeric":
                # Numeric columns can't sensibly be used to group/color a chart (too many
                # distinct values -> hundreds of tiny stacked slivers + a slow, useless
                # continuous color legend). Treat this as "sum this column" instead.
                breakdown_col = None
                st.caption(
                    f"ℹ️ `{breakdown_choice}` is numeric, so it'll be **summed**, not used to split the chart."
                )
                want_sum_choice = st.radio(
                    f"➕ Show the SUM of `{breakdown_choice}` for each month/year?",
                    options=["No", "Yes"],
                    index=0,
                    key=f"want_sum_{selected_table}_{breakdown_choice}",
                    horizontal=True,
                )
                if want_sum_choice == "Yes":
                    sum_columns = [breakdown_choice]
                    crores_choice = st.radio(
                        f"💰 Display the sum of `{breakdown_choice}` in Crores (÷ 1,00,00,000)?",
                        options=["No", "Yes"],
                        index=0,
                        key=f"crores_choice_{selected_table}_{breakdown_choice}",
                        horizontal=True,
                    )
                    show_in_crores = crores_choice == "Yes"
            else:
                # Categorical/text column -> normal group/color breakdown, same as before.
                breakdown_col = breakdown_choice

            try:
                available_ym_summary = get_available_year_months(engine, selected_table, summary_date_col)
            except SQLAlchemyError as exc:
                st.error(f"Failed to fetch available months: {exc}")
                available_ym_summary = []
            ym_labels_summary = {f"{MONTH_NAMES[mo - 1]} {yr}": (yr, mo) for yr, mo in available_ym_summary}
            focus_labels = st.multiselect(
                "Focus on specific month(s) — leave empty to see every month",
                options=list(ym_labels_summary.keys()),
                key=f"focus_months_{selected_table}_{summary_date_col}",
            )
            focus_year_months = [ym_labels_summary[lbl] for lbl in focus_labels]

            try:
                available_years_summary = get_available_years(engine, selected_table, summary_date_col)
            except SQLAlchemyError as exc:
                st.error(f"Failed to fetch available years: {exc}")
                available_years_summary = []
            focus_years_summary = st.multiselect(
                "Focus on specific year(s) — leave empty to see every year",
                options=available_years_summary,
                key=f"focus_years_summary_{selected_table}_{summary_date_col}",
            )
            if focus_year_months and focus_years_summary:
                st.caption(
                    "⚠️ Both a specific-month filter and a year filter are set — "
                    "the specific-month filter takes priority below. Clear it if you "
                    "just want the whole year(s) you picked."
                )

            if breakdown_col:
                try:
                    breakdown_values = get_distinct_values(engine, selected_table, breakdown_col)
                except SQLAlchemyError as exc:
                    st.error(f"Failed to fetch values for `{breakdown_col}`: {exc}")
                    breakdown_values = []
                focus_categories = st.multiselect(
                    f"Focus on specific {breakdown_col} value(s) — e.g. Cash — leave empty for all",
                    options=breakdown_values,
                    key=f"focus_categories_{selected_table}_{breakdown_col}",
                )

            chart_res_scale = st.select_slider(
                "🖼️ Chart clarity (export resolution)",
                options=[1, 2, 3, 4],
                value=3,
                key="chart_res_scale",
                help=(
                    "Base chart is 1600×800px. 3× ≈ 4800×2400px and 4× ≈ 6400×3200px — "
                    "both exceed 4K (3840×2160) for crisp downloads."
                ),
            )

            num_chart_splits = st.slider(
                "📊 Split the monthly chart into how many charts?",
                min_value=1,
                max_value=12,
                value=1,
                key="num_chart_splits",
                help=(
                    "If you have many months, one chart can get congested. "
                    "Raise this to break it into that many smaller charts "
                    "(e.g. 3 → roughly a third of the months per chart)."
                ),
            )

    st.divider()
    st.subheader("4️⃣ Summary table Analysis (optional)")
    show_pivot_table = st.checkbox(
        "Show monthly cumulative summary ",
        value=False,
        key=f"show_pivot_table_{selected_table}",
        help=(
            "Build a 4-quadrant Excel-style PivotTable: pick Filters, Rows, "
            "Columns, and a Value + aggregation, just like Excel's PivotTable."
        ),
    )

    pivot_filter_cols: list[str] = []
    pivot_filter_values: dict[str, list[Any]] = {}
    pivot_row_cols: list[str] = []
    pivot_col_cols: list[str] = []
    pivot_value_cols: list[str] = []
    pivot_agg_func: str = "sum"
    pivot_show_in_crores: bool = False
    pivot_right_total_mode: str = "grand_total"
    pivot_run_clicked = False
    pivot_is_built = st.session_state.get(f"pivot_built_{selected_table}", False)

    if show_pivot_table:
        numeric_cols = [c for c in column_names if column_kind_map.get(c) == "numeric"]

        st.markdown("**🔍 Filters**")
        pivot_filter_cols = st.multiselect(
            "Column(s) to filter the dataset before pivoting",
            options=column_names,
            key=f"pivot_filter_cols_{selected_table}",
        )
        for fc in pivot_filter_cols:
            try:
                fc_values = get_distinct_values(engine, selected_table, fc)
            except SQLAlchemyError as exc:
                st.error(f"Failed to fetch values for `{fc}`: {exc}")
                fc_values = []
            pivot_filter_values[fc] = st.multiselect(
                f"↳ Value(s) of `{fc}` — leave empty to include all",
                options=fc_values,
                key=f"pivot_filter_vals_{selected_table}_{fc}",
            )

        st.markdown("**≡ Rows**")
        pivot_row_cols = st.multiselect(
            "Column(s) to group vertically (e.g. ZONE, BRANCH)",
            options=[c for c in column_names if c not in pivot_filter_cols],
            key=f"pivot_row_cols_{selected_table}",
        )

        st.markdown("**|||| Columns**")
        pivot_col_cols = st.multiselect(
            "Column(s) to split horizontally across headers (e.g. STAGE_ECL, STATUS)",
            options=[c for c in column_names if c not in pivot_row_cols],
            key=f"pivot_col_cols_{selected_table}",
        )

        if len(pivot_col_cols) >= 2:
            deepest_label = pivot_col_cols[-1]
            outer_labels = ", ".join(pivot_col_cols[:-1])
            right_total_help = {
                "deepest_field_total": (
                    f"One '<value> Total' column per {deepest_label} (e.g. one per month), "
                    f"each summed across every {outer_labels} — the single overall Grand Total "
                    "column is dropped."
                ),
                "grand_total": "The original single 'Grand Total' column, summing everything together.",
                "both": (
                    f"The per-{deepest_label} Total columns, plus one overall Grand Total "
                    "at the very end."
                ),
            }
            pivot_right_total_mode = st.radio(
                f"↳ Rightmost total column(s) for `{deepest_label}`",
                options=["deepest_field_total", "grand_total", "both"],
                format_func=lambda k: {
                    "deepest_field_total": f"Per-{deepest_label} totals (new)",
                    "grand_total": "Single overall Grand Total (original)",
                    "both": "Both",
                }[k],
                index=0,
                key=f"pivot_right_total_mode_{selected_table}",
                help=right_total_help["deepest_field_total"],
            )
            st.caption(right_total_help[pivot_right_total_mode])
        else:
            pivot_right_total_mode = "grand_total"

        st.markdown("**∑ Values & Aggregation**")
        pivot_value_cols = st.multiselect(
            "Column(s) to aggregate — pick more than one to get each as its own set of "
            "columns (e.g. \"Sum of EMI_Mar23\", \"Sum of EMI_Apr23\"), just like Excel's "
            "multiple Values fields",
            options=column_names,
            key=f"pivot_value_cols_{selected_table}",
        )
        pivot_agg_func = st.selectbox(
            "Aggregation function (applied to every Values column picked above)",
            options=PIVOT_AGG_FUNCS,
            key=f"pivot_agg_func_{selected_table}",
        )

        if pivot_value_cols and pivot_agg_func != "count":
            crores_pivot_choice = st.radio(
                "💰 Display value(s) in Crores (÷ 1,00,00,000) in the Summary Table?",
                options=["No", "Yes"],
                index=0,
                key=f"pivot_crores_{selected_table}",
                horizontal=True,
            )
            pivot_show_in_crores = crores_pivot_choice == "Yes"

        st.caption("A Grand Total row is always added. Column totals follow your choice above (if shown).")
        pivot_run_clicked = st.button(
            "🧮 Build Summary", type="primary", use_container_width=True, key=f"pivot_run_{selected_table}"
        )
        if pivot_run_clicked:
            # A button's True value only lasts for the one rerun it was clicked
            # on — the very next rerun (e.g. touching a widget in the ratio
            # calculator below, or tweaking Rows/Columns) it reports False
            # again. Persist "this table's pivot has been built" separately so
            # the whole Summary Table doesn't vanish and reappear stale on
            # every unrelated interaction; it now stays live and always
            # reflects the *current* Rows/Columns/Values choices.
            st.session_state[f"pivot_built_{selected_table}"] = True
        pivot_is_built = st.session_state.get(f"pivot_built_{selected_table}", False)

    st.divider()
    run_clicked = st.button("🚀 Generate Report", type="primary", use_container_width=True)

# --------------------------------------------------------------------------- #
# Main UI - Tabs Layout
# --------------------------------------------------------------------------- #
_access_level = st.session_state.get("auth_access_level", "user")
_can_use_sql_tab = _access_level in ("admin", "analyst")

_tab_labels = ["📊 Dynamic Report Builder", "🧮 Summary Table"]
if _can_use_sql_tab:
    _tab_labels.append("📤 Upload Data")
    _tab_labels.append("🧑‍💻 Custom SQL Query")
if _access_level == "admin":
    _tab_labels.append("🛡️ Admin")

_tabs = st.tabs(_tab_labels)
tab_report, tab_pivot = _tabs[0], _tabs[1]
tab_upload = _tabs[2] if _can_use_sql_tab else None
tab_sql = _tabs[3] if _can_use_sql_tab else None
tab_admin = _tabs[-1] if _access_level == "admin" else None

if tab_upload is not None:
    with tab_upload:
        render_upload_tab(engine)

if tab_admin is not None:
    with tab_admin:
        render_admin_panel(engine)

with tab_report:
    # --------------------------------------------------------------------------- #
    # Monthly Cumulative Summary (independent of the report section below)
    # --------------------------------------------------------------------------- #
    if show_monthly_summary and summary_date_col:
        st.subheader("🗓️ Monthly Cumulative Summary")
        st.caption(
            f"Every record is counted under the calendar month of `{summary_date_col}` "
            "— e.g. a record dated 01-03-2024 is counted as a **March** record, regardless of the day."
        )

        try:
            with st.spinner("Building monthly summary..."):
                summary_df = get_monthly_summary(
                    engine, selected_table, summary_date_col, breakdown_col, tuple(sum_columns)
                )
        except SQLAlchemyError as exc:
            st.error(f"❌ Failed to build monthly summary: {exc}")
            summary_df = pd.DataFrame()
        except Exception as exc:  # noqa: BLE001
            st.error(f"❌ Unexpected error while building monthly summary: {exc}")
            summary_df = pd.DataFrame()

        if summary_df.empty:
            st.info("No data available to summarize for this date column.")
        else:
            summary_df["month_label"] = summary_df.apply(
                lambda r: f"{MONTH_NAMES[int(r['mo']) - 1]} {int(r['yr'])}", axis=1
            )
            summary_df = summary_df.sort_values(["yr", "mo"])

            # Apply optional focus filters chosen in the sidebar.
            # If specific month(s) are picked, that's the most precise filter and wins.
            # Otherwise, fall back to year-only focus so picking "2023" here also
            # narrows this monthly view down to just 2023 (previously this only
            # affected the Yearly section below, which made the monthly chart/table/
            # CSV look "unfiltered" even after picking a year).
            display_df = summary_df.copy()
            if focus_year_months:
                focus_set = {(int(yr), int(mo)) for yr, mo in focus_year_months}
                display_df = display_df[
                    display_df.apply(lambda r: (int(r["yr"]), int(r["mo"])) in focus_set, axis=1)
                ]
            elif focus_years_summary:
                focus_years_set = {int(y) for y in focus_years_summary}
                display_df = display_df[display_df["yr"].astype(int).isin(focus_years_set)]
            if breakdown_col and focus_categories:
                display_df = display_df[display_df[breakdown_col].isin(focus_categories)]

            if display_df.empty:
                st.warning(
                    "No records match the selected month/category focus. "
                    f"Active filters — Month(s): {focus_labels or 'none'}, "
                    f"Year(s): {focus_years_summary or 'none'}, "
                    f"{breakdown_col or 'category'}: {focus_categories or 'none'}. "
                    "Clear one of these in the sidebar to see results."
                )
            elif breakdown_col:
                # --- Grouped comparison chart: month x category (with value labels) ---
                render_split_bar_charts(
                    display_df,
                    x_col="month_label",
                    y_col="record_count",
                    color_col=breakdown_col,
                    title=f"Records per month by {breakdown_col}",
                    filename_prefix=f"{selected_table}_monthly_{breakdown_col}",
                    key_suffix="monthly_breakdown",
                    res_scale=chart_res_scale,
                    num_splits=num_chart_splits,
                )

                # --- Excel-style pivot: rows = month, columns = category ---
                pivot_df = display_df.pivot_table(
                    index="month_label",
                    columns=breakdown_col,
                    values="record_count",
                    aggfunc="sum",
                    fill_value=0,
                    margins=True,
                    margins_name="Total",
                )
                month_order = [
                    m for m in display_df.sort_values(["yr", "mo"])["month_label"].unique() if m in pivot_df.index
                ]
                if "Total" in pivot_df.index:
                    month_order.append("Total")
                pivot_df = pivot_df.reindex(month_order)
                st.dataframe(pivot_df, use_container_width=True)

                summary_csv_buffer = io.StringIO()
                pivot_df.to_csv(summary_csv_buffer)
                st.download_button(
                    "⬇️ Download Monthly Breakdown CSV",
                    data=summary_csv_buffer.getvalue(),
                    file_name=f"{selected_table}_monthly_{breakdown_col}_summary.csv",
                    mime="text/csv",
                    key="download_monthly_breakdown",
                )

                # --- Plain row-by-row data (Month | category | Count) ---------
                with st.expander("📋 View all rows (Month × " + breakdown_col + ")", expanded=False):
                    flat_df = display_df[["month_label", breakdown_col, "record_count"]].rename(
                        columns={"month_label": "Month", "record_count": "Count"}
                    )
                    st.dataframe(flat_df, use_container_width=True, hide_index=True)
                    flat_csv_buffer = io.StringIO()
                    flat_df.to_csv(flat_csv_buffer, index=False)
                    st.download_button(
                        "⬇️ Download Filtered Rows CSV",
                        data=flat_csv_buffer.getvalue(),
                        file_name=f"{selected_table}_monthly_{breakdown_col}_rows.csv",
                        mime="text/csv",
                        key="download_monthly_breakdown_rows",
                    )
            else:
                # --- Simple month-over-month comparison (with value labels) ---
                render_split_bar_charts(
                    display_df,
                    x_col="month_label",
                    y_col="record_count",
                    color_col=None,
                    title="Records per month",
                    filename_prefix=f"{selected_table}_monthly_summary",
                    key_suffix="monthly_simple",
                    res_scale=chart_res_scale,
                    num_splits=num_chart_splits,
                )

                if focus_year_months:
                    cols = st.columns(min(len(display_df), 4) or 1)
                    for i, (_, row) in enumerate(display_df.iterrows()):
                        cols[i % len(cols)].metric(row["month_label"], f"{int(row['record_count']):,}")

                st.dataframe(
                    display_df[["month_label", "record_count"]].rename(
                        columns={"month_label": "Month", "record_count": "Count"}
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                summary_csv_buffer = io.StringIO()
                display_df[["month_label", "record_count"]].to_csv(summary_csv_buffer, index=False)
                st.download_button(
                    "⬇️ Download Monthly Summary CSV",
                    data=summary_csv_buffer.getvalue(),
                    file_name=f"{selected_table}_monthly_summary.csv",
                    mime="text/csv",
                    key="download_monthly_summary",
                )

            # --- Optional SUM columns (e.g. "emi") per month ------------------- #
            if not display_df.empty and sum_columns:
                unit_note = " (shown in Crores)" if show_in_crores else ""
                for sc in sum_columns:
                    sum_field = f"sum_{sc}"
                    if sum_field not in display_df.columns:
                        continue

                    st.markdown(f"#### 💵 Sum of `{sc}` per month{unit_note}")

                    chart_df = display_df.copy()
                    if show_in_crores:
                        chart_df[sum_field] = chart_df[sum_field] / CRORE
                    y_axis_label = f"Sum of {sc}" + (" (Cr)" if show_in_crores else "")
                    bar_value_format = ",.2f" if show_in_crores else ",.0f"

                    if breakdown_col:
                        render_split_bar_charts(
                            chart_df,
                            x_col="month_label",
                            y_col=sum_field,
                            color_col=breakdown_col,
                            title=f"Sum of {sc} per month by {breakdown_col}{unit_note}",
                            filename_prefix=f"{selected_table}_monthly_sum_{sc}_{breakdown_col}",
                            key_suffix=f"monthly_sum_{sc}_breakdown",
                            res_scale=chart_res_scale,
                            num_splits=num_chart_splits,
                            y_label=y_axis_label,
                            value_format=bar_value_format,
                        )

                        sum_pivot_df = display_df.pivot_table(
                            index="month_label",
                            columns=breakdown_col,
                            values=sum_field,
                            aggfunc="sum",
                            fill_value=0,
                            margins=True,
                            margins_name="Total",
                        )
                        sum_pivot_df = sum_pivot_df.reindex(
                            [m for m in month_order if m in sum_pivot_df.index]
                        )
                        sum_pivot_view = (
                            sum_pivot_df.div(CRORE) if show_in_crores else sum_pivot_df
                        ).round(2)
                        st.dataframe(sum_pivot_view, use_container_width=True)
                        if show_in_crores:
                            st.caption("Table above is in Crores. Downloaded CSV keeps raw (non-Crore) values.")

                        sum_csv_buffer = io.StringIO()
                        sum_pivot_df.to_csv(sum_csv_buffer)  # always raw values in CSV
                        st.download_button(
                            f"⬇️ Download Monthly Sum({sc}) Breakdown CSV",
                            data=sum_csv_buffer.getvalue(),
                            file_name=f"{selected_table}_monthly_sum_{sc}_{breakdown_col}.csv",
                            mime="text/csv",
                            key=f"download_monthly_sum_{sc}_breakdown",
                        )
                    else:
                        render_split_bar_charts(
                            chart_df,
                            x_col="month_label",
                            y_col=sum_field,
                            color_col=None,
                            title=f"Sum of {sc} per month{unit_note}",
                            filename_prefix=f"{selected_table}_monthly_sum_{sc}",
                            key_suffix=f"monthly_sum_{sc}",
                            res_scale=chart_res_scale,
                            num_splits=num_chart_splits,
                            y_label=y_axis_label,
                            value_format=bar_value_format,
                        )

                        if focus_year_months:
                            m_cols = st.columns(min(len(display_df), 4) or 1)
                            for i, (_, row) in enumerate(display_df.iterrows()):
                                m_cols[i % len(m_cols)].metric(
                                    row["month_label"], format_amount(row[sum_field], show_in_crores)
                                )

                        sum_table_df = display_df[["month_label", sum_field]].rename(
                            columns={"month_label": "Month", sum_field: f"Sum({sc})"}
                        )
                        table_view = sum_table_df.copy()
                        if show_in_crores:
                            table_view[f"Sum({sc}) [Cr]"] = series_to_crores(table_view[f"Sum({sc})"]).round(2)
                            table_view = table_view.drop(columns=[f"Sum({sc})"])
                        st.dataframe(table_view, use_container_width=True, hide_index=True)

                        sum_csv_buffer = io.StringIO()
                        sum_table_df.to_csv(sum_csv_buffer, index=False)  # raw values in CSV
                        st.download_button(
                            f"⬇️ Download Monthly Sum({sc}) CSV",
                            data=sum_csv_buffer.getvalue(),
                            file_name=f"{selected_table}_monthly_sum_{sc}.csv",
                            mime="text/csv",
                            key=f"download_monthly_sum_{sc}",
                        )
                    st.divider()
        st.divider()

        # --------------------------------------------------------------------------- #
        # Yearly Cumulative Summary — same data, rolled up to 2023 / 2024 / 2025 ...
        # --------------------------------------------------------------------------- #
        st.subheader("📅 Yearly Cumulative Summary")
        st.caption(
            f"Same data as above, rolled up to the calendar year of `{summary_date_col}` "
            "(e.g. 2023, 2024, 2025)."
        )

        group_cols = ["yr"] + ([breakdown_col] if breakdown_col else [])
        yearly_agg_cols = ["record_count"] + [f"sum_{sc}" for sc in sum_columns if f"sum_{sc}" in summary_df.columns]
        yearly_df = summary_df.groupby(group_cols, as_index=False)[yearly_agg_cols].sum()
        yearly_df["year_label"] = yearly_df["yr"].astype(int).astype(str)
        yearly_df = yearly_df.sort_values("yr")

        yearly_display_df = yearly_df.copy()
        if focus_years_summary:
            focus_years_set = {int(y) for y in focus_years_summary}
            yearly_display_df = yearly_display_df[yearly_display_df["yr"].astype(int).isin(focus_years_set)]
        if breakdown_col and focus_categories:
            yearly_display_df = yearly_display_df[yearly_display_df[breakdown_col].isin(focus_categories)]

        if yearly_display_df.empty:
            st.warning(
                "No records match the selected year/category focus. "
                f"Active filters — Year(s): {focus_years_summary or 'none'}, "
                f"{breakdown_col or 'category'}: {focus_categories or 'none'}. "
                "Clear one of these in the sidebar to see results."
            )
        elif breakdown_col:
            render_split_bar_charts(
                yearly_display_df,
                x_col="year_label",
                y_col="record_count",
                color_col=breakdown_col,
                title=f"Records per year by {breakdown_col}",
                filename_prefix=f"{selected_table}_yearly_{breakdown_col}",
                key_suffix="yearly_breakdown",
                res_scale=chart_res_scale,
                num_splits=num_chart_splits,
            )

            yearly_pivot_df = yearly_display_df.pivot_table(
                index="year_label",
                columns=breakdown_col,
                values="record_count",
                aggfunc="sum",
                fill_value=0,
                margins=True,
                margins_name="Total",
            )
            year_order = [
                y for y in yearly_display_df.sort_values("yr")["year_label"].unique() if y in yearly_pivot_df.index
            ]
            if "Total" in yearly_pivot_df.index:
                year_order.append("Total")
            yearly_pivot_df = yearly_pivot_df.reindex(year_order)
            st.dataframe(yearly_pivot_df, use_container_width=True)

            yearly_csv_buffer = io.StringIO()
            yearly_pivot_df.to_csv(yearly_csv_buffer)
            st.download_button(
                "⬇️ Download Yearly Breakdown CSV",
                data=yearly_csv_buffer.getvalue(),
                file_name=f"{selected_table}_yearly_{breakdown_col}_summary.csv",
                mime="text/csv",
                key="download_yearly_breakdown",
            )

            # --- Plain row-by-row data (Year | category | Count) -------------
            with st.expander("📋 View all rows (Year × " + breakdown_col + ")", expanded=False):
                yearly_flat_df = yearly_display_df[["year_label", breakdown_col, "record_count"]].rename(
                    columns={"year_label": "Year", "record_count": "Count"}
                )
                st.dataframe(yearly_flat_df, use_container_width=True, hide_index=True)
                yearly_flat_csv_buffer = io.StringIO()
                yearly_flat_df.to_csv(yearly_flat_csv_buffer, index=False)
                st.download_button(
                    "⬇️ Download Filtered Rows CSV",
                    data=yearly_flat_csv_buffer.getvalue(),
                    file_name=f"{selected_table}_yearly_{breakdown_col}_rows.csv",
                    mime="text/csv",
                    key="download_yearly_breakdown_rows",
                )
        else:
            render_split_bar_charts(
                yearly_display_df,
                x_col="year_label",
                y_col="record_count",
                color_col=None,
                title="Records per year",
                filename_prefix=f"{selected_table}_yearly_summary",
                key_suffix="yearly_simple",
                res_scale=chart_res_scale,
                num_splits=num_chart_splits,
            )

            if focus_years_summary:
                cols = st.columns(min(len(yearly_display_df), 4) or 1)
                for i, (_, row) in enumerate(yearly_display_df.iterrows()):
                    cols[i % len(cols)].metric(row["year_label"], f"{int(row['record_count']):,}")

            st.dataframe(
                yearly_display_df[["year_label", "record_count"]].rename(
                    columns={"year_label": "Year", "record_count": "Count"}
                ),
                use_container_width=True,
                hide_index=True,
            )

            yearly_csv_buffer = io.StringIO()
            yearly_display_df[["year_label", "record_count"]].to_csv(yearly_csv_buffer, index=False)
            st.download_button(
                "⬇️ Download Yearly Summary CSV",
                data=yearly_csv_buffer.getvalue(),
                file_name=f"{selected_table}_yearly_summary.csv",
                mime="text/csv",
                key="download_yearly_summary",
            )

        # --- Optional SUM columns (e.g. "emi") per year ------------------- #
        if not yearly_display_df.empty and sum_columns:
            unit_note = " (shown in Crores)" if show_in_crores else ""
            for sc in sum_columns:
                sum_field = f"sum_{sc}"
                if sum_field not in yearly_display_df.columns:
                    continue

                st.markdown(f"#### 💵 Sum of `{sc}` per year{unit_note}")

                yearly_chart_df = yearly_display_df.copy()
                if show_in_crores:
                    yearly_chart_df[sum_field] = yearly_chart_df[sum_field] / CRORE
                yearly_y_axis_label = f"Sum of {sc}" + (" (Cr)" if show_in_crores else "")
                yearly_bar_value_format = ",.2f" if show_in_crores else ",.0f"

                if breakdown_col:
                    render_split_bar_charts(
                        yearly_chart_df,
                        x_col="year_label",
                        y_col=sum_field,
                        color_col=breakdown_col,
                        title=f"Sum of {sc} per year by {breakdown_col}{unit_note}",
                        filename_prefix=f"{selected_table}_yearly_sum_{sc}_{breakdown_col}",
                        key_suffix=f"yearly_sum_{sc}_breakdown",
                        res_scale=chart_res_scale,
                        num_splits=num_chart_splits,
                        y_label=yearly_y_axis_label,
                        value_format=yearly_bar_value_format,
                    )

                    yearly_sum_pivot_df = yearly_display_df.pivot_table(
                        index="year_label",
                        columns=breakdown_col,
                        values=sum_field,
                        aggfunc="sum",
                        fill_value=0,
                        margins=True,
                        margins_name="Total",
                    )
                    yearly_sum_pivot_df = yearly_sum_pivot_df.reindex(
                        [y for y in year_order if y in yearly_sum_pivot_df.index]
                    )
                    yearly_sum_pivot_view = (
                        yearly_sum_pivot_df.div(CRORE) if show_in_crores else yearly_sum_pivot_df
                    ).round(2)
                    st.dataframe(yearly_sum_pivot_view, use_container_width=True)
                    if show_in_crores:
                        st.caption("Table above is in Crores. Downloaded CSV keeps raw (non-Crore) values.")

                    yearly_sum_csv_buffer = io.StringIO()
                    yearly_sum_pivot_df.to_csv(yearly_sum_csv_buffer)  # raw values
                    st.download_button(
                        f"⬇️ Download Yearly Sum({sc}) Breakdown CSV",
                        data=yearly_sum_csv_buffer.getvalue(),
                        file_name=f"{selected_table}_yearly_sum_{sc}_{breakdown_col}.csv",
                        mime="text/csv",
                        key=f"download_yearly_sum_{sc}_breakdown",
                    )
                else:
                    render_split_bar_charts(
                        yearly_chart_df,
                        x_col="year_label",
                        y_col=sum_field,
                        color_col=None,
                        title=f"Sum of {sc} per year{unit_note}",
                        filename_prefix=f"{selected_table}_yearly_sum_{sc}",
                        key_suffix=f"yearly_sum_{sc}",
                        res_scale=chart_res_scale,
                        num_splits=num_chart_splits,
                        y_label=yearly_y_axis_label,
                        value_format=yearly_bar_value_format,
                    )

                    if focus_years_summary:
                        y_cols = st.columns(min(len(yearly_display_df), 4) or 1)
                        for i, (_, row) in enumerate(yearly_display_df.iterrows()):
                            y_cols[i % len(y_cols)].metric(
                                row["year_label"], format_amount(row[sum_field], show_in_crores)
                            )

                    yearly_sum_table_df = yearly_display_df[["year_label", sum_field]].rename(
                        columns={"year_label": "Year", sum_field: f"Sum({sc})"}
                    )
                    yearly_table_view = yearly_sum_table_df.copy()
                    if show_in_crores:
                        yearly_table_view[f"Sum({sc}) [Cr]"] = series_to_crores(
                            yearly_table_view[f"Sum({sc})"]
                        ).round(2)
                        yearly_table_view = yearly_table_view.drop(columns=[f"Sum({sc})"])
                    st.dataframe(yearly_table_view, use_container_width=True, hide_index=True)

                    yearly_sum_csv_buffer = io.StringIO()
                    yearly_sum_table_df.to_csv(yearly_sum_csv_buffer, index=False)  # raw values
                    st.download_button(
                        f"⬇️ Download Yearly Sum({sc}) CSV",
                        data=yearly_sum_csv_buffer.getvalue(),
                        file_name=f"{selected_table}_yearly_sum_{sc}.csv",
                        mime="text/csv",
                        key=f"download_yearly_sum_{sc}",
                    )
                st.divider()
        st.divider()

    # --------------------------------------------------------------------------- #
    # Main panel - results
    # --------------------------------------------------------------------------- #
    if not output_columns:
        st.info("👈 Select at least one output column in the sidebar to build a report.")
    elif not run_clicked:
        st.info("👈 Configure your filters and click **Generate Report** to run the query.")
    else:
        query, params = build_query(selected_table, ref_column, output_columns, ref_kind, filter_payload)

        with st.expander("🔍 View generated SQL", expanded=False):
            st.code(query, language="sql")
            if params:
                st.caption("Parameters:")
                st.json({k: str(v) for k, v in params.items()})

        try:
            with st.spinner("Running query..."):
                df = run_query(engine, query, params)
                
            st.success(f"Query returned **{len(df):,}** row(s).")

            # --- KPI cards for numeric output columns ---
            numeric_output_cols = [
                c for c in output_columns
                if c in df.columns and pd.api.types.is_numeric_dtype(df[c])
            ]

            if numeric_output_cols:
                st.subheader("📈 KPI Summary")
                if show_in_crores:
                    st.caption("💰 Sum/Average shown in Crores (÷1,00,00,000).")
                for col in numeric_output_cols:
                    series = df[col].dropna()
                    c1, c2, c3 = st.columns(3)
                    c1.metric(f"Total Sum — {col}", format_amount(series.sum(), show_in_crores))
                    c2.metric(
                        f"Average — {col}",
                        format_amount(series.mean(), show_in_crores) if len(series) else format_amount(0, show_in_crores),
                    )
                    c3.metric(f"Count — {col}", f"{series.count():,}")
                st.divider()

            # --- Data table ---
            st.subheader("📋 Results")
            st.dataframe(df, use_container_width=True, height=450)
            st.caption(f"Total rows: **{len(df):,}** | Columns: {', '.join(df.columns)}")

            # --- Download button ---
            csv_buffer = io.StringIO()
            df.to_csv(csv_buffer, index=False)
            st.download_button(
                label="⬇️ Download CSV",
                data=csv_buffer.getvalue(),
                file_name=f"{selected_table}_report.csv",
                mime="text/csv",
                use_container_width=True,
            )
        except SQLAlchemyError as exc:
            st.error(f"❌ Query failed: {exc}")
        except Exception as exc:  # noqa: BLE001
            st.error(f"❌ Unexpected error while running the query: {exc}")

with tab_pivot:
    st.subheader("🧮 Quadrant Summary Table")

    if not show_pivot_table:
        st.info(
            "👈 Enable **'Show monthly cumulative summary / Pivot Table'** in the sidebar "
            "(section 4️⃣) to build a pivot table."
        )
    elif not pivot_row_cols:
        st.info("👈 Choose at least one **Rows** column in the sidebar to build a pivot table.")
    elif not pivot_value_cols:
        st.info("👈 Choose at least one numeric **Values** column in the sidebar to aggregate.")
    elif not pivot_is_built:
        st.info("👈 Configure Filters / Rows / Columns / Values in the sidebar, then click **Build Summary** once. It'll then stay live and update automatically as you change your selections or use the calculator below.")
    else:
        try:
            filter_conditions = tuple(
                (fc, tuple(pivot_filter_values.get(fc, []))) for fc in pivot_filter_cols
            )

            with st.spinner("Aggregating data in the database..."):
                pivot_source_dfs = {
                    vc: fetch_pivot_source_data(
                        engine,
                        selected_table,
                        tuple(pivot_row_cols),
                        tuple(pivot_col_cols),
                        vc,
                        filter_conditions,
                        column_kind_map,
                    )
                    for vc in pivot_value_cols
                }
            pivot_source_df = pivot_source_dfs[pivot_value_cols[0]]  # first Values column -- used below by the Chart, Field Efficiency Table, and DPD buckets

            if any(df.empty for df in pivot_source_dfs.values()):
                st.warning("No rows match the selected filters — nothing to pivot.")
            else:
                # Flat/chart view uses each date field at its exact-date leaf
                # granularity only (matching the original single-level
                # behaviour) -- the Summary Matrix below is the one that
                # shows the full Month > Date nesting.
                flat_row_cols = [_date_field_leaf(f, column_kind_map) for f in pivot_row_cols]
                flat_col_cols = [_date_field_leaf(f, column_kind_map) for f in pivot_col_cols]

                with st.spinner("Building pivot table..."):
                    pivot_df = build_pivot_table(
                        pivot_source_df, flat_row_cols, flat_col_cols, pivot_agg_func
                    )

                st.success(
                    f"Pivot built from **{len(pivot_source_df):,}** aggregated group(s) — "
                    f"**{pivot_agg_func}({', '.join(pivot_value_cols)})** by **{', '.join(pivot_row_cols)}**"
                    + (f" × **{', '.join(pivot_col_cols)}**" if pivot_col_cols else "")
                )
                st.caption(
                    "Aggregated server-side in PostgreSQL (GROUP BY, with date fields auto-bucketed by month) — "
                    "every view below is built from this small aggregated table, not from the raw rows."
                )

                st.markdown("#### 📋 Summary Matrix")
                if len(pivot_value_cols) == 1:
                    display_pivot_df = render_excel_style_pivot_table(
                        pivot_source_df,
                        pivot_row_cols,
                        pivot_col_cols,
                        pivot_value_cols[0],
                        pivot_agg_func,
                        column_kind_map,
                        pivot_show_in_crores,
                        pivot_right_total_mode,
                    )
                else:
                    display_pivot_df = render_multi_value_pivot_table(
                        pivot_source_dfs,
                        pivot_row_cols,
                        pivot_col_cols,
                        pivot_value_cols,
                        pivot_agg_func,
                        column_kind_map,
                        pivot_show_in_crores,
                        pivot_right_total_mode,
                    )

                # --- Optional: cumulative DPD buckets (1+ / 30+ / 90+) ------ #
                # Opt-in since most pivots built with this generic tool won't
                # be about DPD slabs at all -- only offered when one of the
                # chosen Columns fields actually looks like a DPD slab field.
                # Kept to the single-Values-column case: with 2+ Values
                # columns the header already has an extra outer level, which
                # would shift where the DPD slab level sits.
                if pivot_col_cols and len(pivot_value_cols) == 1:
                    add_dpd_buckets = st.checkbox(
                        "➕ Add cumulative DPD buckets (1+ / 30+ / 90+) to the Summary Matrix",
                        value=False,
                        key=f"add_dpd_buckets_{selected_table}",
                        help=(
                            "Sums the raw DPD slab columns (e.g. '1-29', '30-59', '60-89', '90+') "
                            "into cumulative macro-buckets: 1+ = all delinquent slabs, "
                            "30+ = 30-59/60-89/90+, 90+ = the 90+ slab alone. Requires one of "
                            "your Columns fields to hold the raw DPD slab labels, and — if you've "
                            "picked 2+ Columns fields — that field must be listed FIRST."
                        ),
                    )
                    if add_dpd_buckets:
                        dpd_field_choice = st.selectbox(
                            "Which Columns field holds the DPD slab?",
                            options=pivot_col_cols,
                            index=0,
                            key=f"dpd_field_choice_{selected_table}",
                        )
                        try:
                            slab_level = _effective_col_level(pivot_col_cols, dpd_field_choice, column_kind_map)
                            display_pivot_df = add_dpd_buckets_to_excel_pivot(
                                display_pivot_df, slab_level, only_show_buckets=True
                            )
                            st.markdown("##### ➕ Summary Matrix — Cumulative DPD Buckets (1+ / 30+ / 90+)")
                            num_fmt = "{:,.0f}" if pivot_agg_func == "count" else "{:,.2f}"
                            st.dataframe(
                                display_pivot_df.style.format(num_fmt),
                                use_container_width=True,
                            )
                        except ValueError as exc:
                            st.warning(f"⚠️ Couldn't add DPD buckets: {exc}")
                elif pivot_col_cols and len(pivot_value_cols) > 1:
                    st.caption(
                        "ℹ️ DPD bucket columns (1+/30+/90+) are only offered when exactly one "
                        "Values column is selected."
                    )

                available_efficiency_fields = list(dict.fromkeys([*pivot_row_cols, *pivot_col_cols]))
                if available_efficiency_fields:
                    st.markdown("#### 🎯 Field Efficiency Table (% of a chosen field's group total)")
                    st.caption(
                        "Pick any one of your selected Rows/Columns fields below — the table will show "
                        "each cell as that field's % share of the total across its own group, holding "
                        "every other selected field fixed. Works the same way whether you point it at a "
                        "Rows field or a Columns field."
                        + (
                            " Shown as one column block per Values column, side by side."
                            if len(pivot_value_cols) > 1 else ""
                        )
                    )
                    normalize_field = st.selectbox(
                        "Normalize as % across:",
                        options=available_efficiency_fields,
                        key=f"pivot_normalize_field_{selected_table}",
                    )
                    if len(pivot_value_cols) == 1:
                        efficiency_display_df = render_field_efficiency_table(
                            pivot_source_df,
                            pivot_row_cols,
                            pivot_col_cols,
                            normalize_field,
                            pivot_value_cols[0],
                            pivot_agg_func,
                            column_kind_map,
                        )
                    else:
                        efficiency_display_df = render_multi_value_field_efficiency_table(
                            pivot_source_dfs,
                            pivot_row_cols,
                            pivot_col_cols,
                            normalize_field,
                            pivot_value_cols,
                            pivot_agg_func,
                            column_kind_map,
                        )

                    # --- Optional: cumulative DPD buckets on the % table too --- #
                    # Mathematically valid ONLY when normalizing across the same
                    # field that holds the DPD slabs -- every slab's % share then
                    # already shares one common denominator (the group total), so
                    # summing them into 1+/30+/90+ is exactly as correct as
                    # summing the raw values would be. Kept to the single-Values-
                    # column case: with 2+ Values columns the header already has
                    # an extra outer level, which would shift where the DPD slab
                    # level sits (same reasoning as the Summary Matrix version).
                    if pivot_col_cols and normalize_field in pivot_col_cols and len(pivot_value_cols) == 1:
                        add_dpd_buckets_eff = st.checkbox(
                            "➕ Add cumulative DPD buckets (1+ / 30+ / 90+) to the Field Efficiency Table",
                            value=False,
                            key=f"add_dpd_buckets_eff_{selected_table}",
                            help=(
                                "Only offered here because you're normalizing across the same field "
                                "that holds your DPD slab labels -- each slab's % already divides by "
                                "the same group total, so 1+/30+/90+ can be built by summing those "
                                "% columns directly, same as the Summary Matrix version above."
                            ),
                        )
                        if add_dpd_buckets_eff:
                            try:
                                eff_slab_level = _effective_col_level(pivot_col_cols, normalize_field, column_kind_map)
                                efficiency_display_df = add_dpd_buckets_to_excel_pivot(
                                    efficiency_display_df, eff_slab_level, only_show_buckets=True
                                )
                                st.markdown("##### ➕ Field Efficiency Table — Cumulative DPD Buckets (1+ / 30+ / 90+)")
                                st.dataframe(
                                    efficiency_display_df.style.format("{:,.2f}%"),
                                    use_container_width=True,
                                )
                            except ValueError as exc:
                                st.warning(f"⚠️ Couldn't add DPD buckets: {exc}")
                    elif pivot_col_cols and normalize_field in pivot_col_cols and len(pivot_value_cols) > 1:
                        st.caption(
                            "ℹ️ DPD bucket columns (1+/30+/90+) on this table are only offered when "
                            "exactly one Values column is selected."
                        )

                    # --- Optional: isolate one status value (e.g. only BOUNCED) ---
                    # Slicing happens AFTER the % table above was already computed
                    # over the full dataset, so BOUNCED% is still correctly
                    # BOUNCED / (BOUNCED + CLEARED + ...), never BOUNCED / BOUNCED.
                    status_values = [
                        str(v)
                        for v in pd.unique(
                            pivot_source_df[_date_field_leaf(normalize_field, column_kind_map)].dropna()
                        )
                    ]
                    isolate_status = st.selectbox(
                        f"Show only one value of `{normalize_field}` (optional)",
                        options=["(show all)"] + status_values,
                        key=f"pivot_eff_isolate_{selected_table}",
                    )
                    if isolate_status != "(show all)":
                        try:
                            efficiency_display_df = slice_field_efficiency_to_status(
                                efficiency_display_df, normalize_field, isolate_status
                            )
                            st.markdown(f"##### 🎯 {isolate_status} only")
                            st.dataframe(
                                efficiency_display_df.style.format("{:,.2f}%"),
                                use_container_width=True,
                            )
                        except ValueError as exc:
                            st.warning(f"⚠️ Couldn't isolate '{isolate_status}': {exc}")

                    # Total columns in this table are always labelled exactly
                    # "Total" (never "<X> Total"), so they can be found
                    # directly from the current columns -- correct even after
                    # the "isolate one status" slice above changes the shape.
                    eff_col_tuples = (
                        list(efficiency_display_df.columns)
                        if isinstance(efficiency_display_df.columns, pd.MultiIndex)
                        else [(c,) for c in efficiency_display_df.columns]
                    )
                    eff_total_flags = [str(t[-1]).strip() == "Total" for t in eff_col_tuples]

                    efficiency_xlsx_bytes = dataframe_to_formatted_excel_bytes(
                        efficiency_display_df, total_col_flags=eff_total_flags,
                        value_kind="percent", table_style="efficiency", sheet_name="Field Efficiency",
                    )
                    st.download_button(
                        "⬇️ Download Field Efficiency Table Excel",
                        data=efficiency_xlsx_bytes,
                        file_name=f"{selected_table}_pivot_efficiency_{normalize_field}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"download_pivot_efficiency_{selected_table}",
                    )
                else:
                    efficiency_display_df = None

                st.markdown("#### 📈 Chart")
                for vc in pivot_value_cols:
                    vc_pivot_df = pivot_df if vc == pivot_value_cols[0] else build_pivot_table(
                        pivot_source_dfs[vc], flat_row_cols, flat_col_cols, pivot_agg_func
                    )
                    render_pivot_chart(
                        vc_pivot_df,
                        flat_row_cols,
                        flat_col_cols,
                        vc,
                        pivot_agg_func,
                        key_suffix=f"{selected_table}_{vc}",
                        display_row_cols=pivot_row_cols,
                        display_col_cols=pivot_col_cols,
                    )

                dl_col1, dl_col2 = st.columns(2)
                with dl_col1:
                    flat_value_kind = "count" if pivot_agg_func == "count" else "number"
                    pivot_xlsx_bytes = dataframe_to_formatted_excel_bytes(
                        pivot_df, value_kind=flat_value_kind, table_style="matrix", sheet_name="Pivot (flat)"
                    )
                    st.download_button(
                        "⬇️ Download Pivot Excel (flat)",
                        data=pivot_xlsx_bytes,
                        file_name=f"{selected_table}_pivot_summary.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"download_pivot_{selected_table}",
                    )
                    if len(pivot_value_cols) > 1:
                        st.caption(f"Flat file reflects your first Values column, `{pivot_value_cols[0]}`, only.")
                with dl_col2:
                    # Recompute the same subtotal/Total flags used to render the
                    # Summary Matrix on screen (cheap -- build_excel_style_pivot
                    # is cached, so this is an instant cache hit, not fresh work)
                    # so the downloaded file bolds/shades exactly the same cells.
                    if len(pivot_value_cols) == 1:
                        _, matrix_group_flags, matrix_total_flags = build_excel_style_pivot(
                            pivot_source_df, pivot_row_cols, pivot_col_cols, pivot_value_cols[0],
                            pivot_agg_func, column_kind_map, pivot_show_in_crores, pivot_right_total_mode,
                        )
                    else:
                        matrix_group_flags: list[bool] = []
                        matrix_total_flags: list[bool] = []
                        for vc in pivot_value_cols:
                            _, block_group_flags, block_total_flags = build_excel_style_pivot(
                                pivot_source_dfs[vc], pivot_row_cols, pivot_col_cols, vc,
                                pivot_agg_func, column_kind_map, pivot_show_in_crores, pivot_right_total_mode,
                            )
                            if not matrix_group_flags:
                                matrix_group_flags = block_group_flags
                            matrix_total_flags.extend(block_total_flags)

                    matrix_value_kind = "count" if pivot_agg_func == "count" else "number"
                    display_xlsx_bytes = dataframe_to_formatted_excel_bytes(
                        display_pivot_df, matrix_group_flags, matrix_total_flags,
                        value_kind=matrix_value_kind, table_style="matrix", sheet_name="Summary Matrix",
                    )
                    st.download_button(
                        "⬇️ Download Summary Matrix Excel (with subtotals)",
                        data=display_xlsx_bytes,
                        file_name=f"{selected_table}_pivot_excel_style.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"download_pivot_excel_style_{selected_table}",
                    )
        except SQLAlchemyError as exc:
            st.error(f"❌ Failed to fetch data for the pivot table: {exc}")
        except Exception as exc:  # noqa: BLE001
            st.error(f"❌ Unexpected error while building the pivot table: {exc}")

if tab_sql is not None:
    with tab_sql:
        st.subheader("🧑‍💻 Custom SQL Execution")
        st.info("Write and execute standard PostgreSQL queries directly against your database.")
    
        # Text area for the user to type queries
        custom_query = st.text_area("SQL Query", height=200, placeholder='SELECT * FROM "public"."agreements" LIMIT 100;')
    
        if st.button("▶️ Run Custom Query", type="primary"):
            if custom_query.strip():
                try:
                    with st.spinner("Executing query..."):
                        with engine.connect() as conn:
                            # Uses the existing engine to run the custom text query
                            custom_df = pd.read_sql(text(custom_query), conn)
                
                    st.success(f"Query returned **{len(custom_df):,}** row(s).")
                    st.dataframe(custom_df, use_container_width=True, height=400)

                    
                    custom_csv = io.StringIO()
                    custom_df.to_csv(custom_csv, index=False)
                    st.download_button(
                        label="⬇️ Download Custom Results (CSV)",
                        data=custom_csv.getvalue(),
                        file_name="custom_sql_results.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )
                except Exception as exc:
                    st.error(f"❌ Query failed: {exc}")
            else:
                st.warning("Please enter a SQL query first.")